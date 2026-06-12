import streamlit as st
import pandas as pd
import re
import requests
from bs4 import BeautifulSoup
from io import BytesIO, StringIO
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import logging
import zipfile
import random
import time
from datetime import datetime
import json
import copy
import csv
import asyncio
import nest_asyncio
nest_asyncio.apply()
from playwright.async_api import async_playwright
import io

st.title("ARCHE screener")
st.header("ARCHE screener")

st.markdown(
    """
    This app can be used to screen a list of CAS/EC numbers uploaded by the user.  
    It provides the **C&L classification** from the ECHA-CHEM website and the **ED information** from a number of sources.  
    Necessary databases and login credentials can be found on Dropbox: Science/Data searches/ED screener/input databases.

    **The following sources are automatically accessed from the ECHA website:**
    - ED list from PPP  
    - ED assessment list  
    - SVHC and SVHC intent  
    - CoRAP  
    - PACT  

    **The following sources need to be uploaded by the user:**
    - ED list from BPR  
    - Food additives list  
    - Food flavourings listings  
    """
)

# To run locally
# streamlit run "/Users/arche/Documents/Python/ARCHEscreeningTool/ARCHE_ED Screener.py"

uploaded_file = st.file_uploader("Upload Excel file to screen: A column with name CAS containing all CAS/EC numbers to screen in individual rows below should be on the first sheet.", type=["xlsx"])
api_file = st.file_uploader("Upload API key NextSDS (NextSDS API key.txt)", type=["txt"])
file_BPR_ED = st.file_uploader("Upload BPR ED file (xlsx): Upload the ED list from BPR (xlsx)", type=["xlsx"])
file_food_add = st.file_uploader("Upload Food additives list (xlsx)", type=["xlsx"])
file_food_flav = st.file_uploader("Upload Food flavourings list (xlsx)", type=["xlsx"])

# In-memory log stream
if "log_stream" not in st.session_state:
    st.session_state.log_stream = StringIO()
    logging.basicConfig(stream=st.session_state.log_stream, level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s')

# List of headers to cycle through to avoid detection when scraping
user_agents_list = [
    'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.83 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:91.0) Gecko/20100101 Firefox/91.0',
    'Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15A372 Safari/604.1',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_2_3) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15',
    'Mozilla/5.0 (Linux; Android 10; SM-G973F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.78 Mobile Safari/537.36',
    'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:98.0) Gecko/20100101 Firefox/98.0'
]

# Normalize IDs (only valid characters)
def normalize_id(value: str) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    # Replace various dash characters with ASCII hyphen-minus
    s = (s.replace('\u2010', '-')  # hyphen
           .replace('\u2011', '-')  # non-breaking hyphen
           .replace('\u2012', '-')  # figure dash
           .replace('\u2013', '-')  # en dash
           .replace('\u2014', '-')  # em dash
           .replace('\u2212', '-')  # minus sign
        )
    # Keep only digits and hyphens
    s = re.sub(r"[^\d\-]", "", s)
    # Collapse multiple hyphens (defensive)
    s = re.sub(r"-{2,}", "-", s)
    return s

# -------- CAS validation --------
def valid_cas_format(cas: str) -> bool:
    """CAS format: X…X-XX-X where first block has 2–7 digits."""
    return re.fullmatch(r"\d{2,7}-\d{2}-\d", cas) is not None

def valid_cas_checksum(cas: str) -> bool:
    """
    Validates the CAS check digit.
    Check digit Z = sum(digit * position) % 10 over the concatenation of the first two blocks,
    counting positions from rightmost starting at 1.
    """
    try:
        parts = cas.split('-')
        if len(parts) != 3:
            return False
        body = parts[0] + parts[1]
        check = int(parts[2])
        total = sum(int(d) * i for i, d in enumerate(reversed(body), start=1))
        return (total % 10) == check
    except Exception:
        return False

def is_valid_cas(cas: str) -> bool:
    return valid_cas_format(cas) and valid_cas_checksum(cas)

# -------- EC validation --------
def is_valid_ec(ec: str) -> bool:
    """
    Validates EC number structural format: XXX-XXX-X (digits only).
    """
    return re.fullmatch(r"\d{3}-\d{3}-\d", ec) is not None

def download_echa_list(echa_url, user_agents_list,source=""):
    try:
        headers = {'User-Agent': random.choice(user_agents_list)}
        responseECHA = requests.get(echa_url, headers=headers)

        unique_substances = None
        if responseECHA.status_code == 200:
            soupECHA = BeautifulSoup(responseECHA.text, "html.parser")
            small_tag = soupECHA.find("small", class_="search-results")
            if small_tag:
                text = small_tag.get_text(strip=True)
                match = re.search(r"of\s+([\d,]+)\s+results", text)
                if match:
                    unique_substances = match.group(1).replace(",", "")  # always strip comma
            soupECHA.decompose()
        responseECHA.close()
        if not unique_substances:
            logging.info("Could not determine the number of unique substances.")
            return None

        # Data or payload sent with the POST request+
        paramsECHA = {
            "p_p_id": "disslists_WAR_disslistsportlet",
            "p_p_lifecycle": "2",
            "p_p_state": "normal",
            "p_p_mode": "view",
            "p_p_resource_id": "exportResults",
            "p_p_cacheability": "cacheLevelPage"
        }
        dataECHA = {
            "_disslists_WAR_disslistsportlet_formDate": int(round(time.time() * 1000)),
            "_disslists_WAR_disslistsportlet_exportColumns": "name,ecNumber,casNumber,lec_submitter,prc_public_status,prc_conclusion,diss_update_date,dte_first_published",
            "_disslists_WAR_disslistsportlet_orderByCol": "diss_update_date",
            "_disslists_WAR_disslistsportlet_orderByType": "asc",
            "_disslists_WAR_disslistsportlet_searchFormColumns": "prc_public_status,prc_conclusion,lec_submitter,dte_intention,dte_assessment,diss_update_date",
            "_disslists_WAR_disslistsportlet_searchFormElements": "DROP_DOWN,DROP_DOWN,DROP_DOWN,DATE_PICKER,DATE_PICKER,DATE_PICKER",
            "_disslists_WAR_disslistsportlet_total": unique_substances,
            "_disslists_WAR_disslistsportlet_exportType": "xls"
        }
        headersEDass = {
            "User-Agent": random.choice(user_agents_list),
            "Content-Type": "application/x-www-form-urlencoded",
        }

        # Adjust based on source
        if source == "PACT":
            paramsECHA["p_p_id"] = "disslists_WAR_disslistsportlet"
            dataECHA = {
                "_disspact_WAR_disspactportlet_formDate": int(round(time.time() * 1000)),
                "_disspact_WAR_disspactportlet_exportColumns": "name,ecNumber,casNumber,DISLIST_CORAP,DISLIST_PBT,DISLIST_DOSSIER_EVALUATION,DISLIST_ED,DISLIST_ARN,DISLIST_ROI_CLH,DISLIST_ROI_SVHC,DISLIST_ANX_14_RECOMMENDATION,DISLIST_ROI_RESTRICTION",
                "_disspact_WAR_disspactportlet_exportDislistsColumns": "DISLIST_CORAP,DISLIST_PBT,DISLIST_DOSSIER_EVALUATION,DISLIST_ED,DISLIST_ARN,DISLIST_ROI_CLH,DISLIST_ROI_SVHC,DISLIST_ANX_14_RECOMMENDATION,DISLIST_ROI_RESTRICTION",
                "_disspact_WAR_disspactportlet_orderByCol": "name",
                "_disspact_WAR_disspactportlet_orderByType": "asc",
                "_disspact_WAR_disspactportlet_orderedSearchableShowListColumns": "DISLIST_PBT_diss_update_date,processes,DISLIST_PBT_diss_concern",
                "_disspact_WAR_disspactportlet_orderedSearchableShowListElements": "DATE_PICKER,MULTI_VALUE,INPUT_TEXT",
                "_disspact_WAR_disspactportlet_orderedSearchableShowListProcessColumns": "PACT,PACT,PACT",
                "_disspact_WAR_disspactportlet_multiValueSearchOperatorprocesses": "AND",
                "_disspact_WAR_disspactportlet_total": unique_substances,
                "_disspact_WAR_disspactportlet_exportType": "xls"
            }
        if source == "SVHC":
            dataECHA["_disslists_WAR_disslistsportlet_exportColumns"] = "name,ecNumber,casNumber,haz_detailed_concern,dte_inclusion,doc_cat_decision,doc_cat_iuclid_dossier,doc_cat_supdoc,doc_cat_rcom,prc_external_remarks",
        if source == "SVHCintent":
            dataECHA["_disslists_WAR_disslistsportlet_exportColumns"] = "name,ecNumber,casNumber,sid_other_info_external,sid_avi_index_no,prc_public_status,dte_intention,sbm_expected_submission,sbm_first_submission,dte_withdrawn,lec_submitter,prc_external_remarks,haz_detailed_concern,dte_public_consult_start,dte_public_consult_deadline,doc_cat_report,doc_cat_rcom,prc_msc_agreement_year,doc_cat_agreement,dte_adoption,doc_cat_supdoc,doc_cat_opinion,dte_opinion,doc_cat_minor_opinion,dte_inclusion,diss_update_date,dte_first_published",
        if source == "CoRAP":
            dataECHA["_disslists_WAR_disslistsportlet_exportColumns"] = "name,ecNumber,casNumber,cnt_country,prc_evaluation_year,lec_submitter,haz_detailed_concern,cse_public_lifecycle,diss_update_date,doc_cat_decision,doc_cat_conclusion,doc_cat_justification,dte_corap_publication,lec_contact_address,lec_organization_name,lec_remarks,prc_appeal_link,prc_external_remarks,diss_concern,relevance,dte_first_published",
        if source == "ARN":
            dataECHA["_disslists_WAR_disslistsportlet_exportColumns"] = (
                "name,ecNumber,casNumber,cse_trigger_categories,prc_public_status,"
                "prc_followup_activity,prc_regulatory_hypothesis,dte_intention,"
                "dte_conclusion_published,doc_cat_conclusion,doc_cat_assessm,group_name,"
                "prc_external_remarks,lec_submitter,lec_organization_name,lec_email,"
                "lec_phone,lec_contact_address,cnt_country,diss_update_date,dte_first_published"
            )
            dataECHA["_disslists_WAR_disslistsportlet_orderByCol"] = "dte_conclusion_published"
            dataECHA["_disslists_WAR_disslistsportlet_orderByType"] = "desc"
            dataECHA["_disslists_WAR_disslistsportlet_searchFormColumns"] = (
                "prc_public_status,lec_submitter,cse_trigger_categories,dte_intention,"
                "prc_followup_activity,dte_conclusion_published,prc_regulatory_hypothesis,"
                "diss_update_date,group_name"
            )
            dataECHA["_disslists_WAR_disslistsportlet_searchFormElements"] = (
                "DROP_DOWN,DROP_DOWN,DROP_DOWN,DATE_PICKER,DROP_DOWN,"
                "DATE_PICKER,MULTI_VALUE,DATE_PICKER,INPUT_TEXT"
            )
            dataECHA["_disslists_WAR_disslistsportlet_substance_identifier_field_key"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_prc_public_status"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_lec_submitter"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_cse_trigger_categories"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_dte_intentionFrom"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_dte_intentionTo"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_prc_followup_activity"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_dte_conclusion_publishedFrom"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_dte_conclusion_publishedTo"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_multiValueSearchOperatorprc_regulatory_hypothesis"] = "AND"
            dataECHA["_disslists_WAR_disslistsportlet_diss_update_dateFrom"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_diss_update_dateTo"] = ""
            dataECHA["_disslists_WAR_disslistsportlet_group_name"] = ""

        responseECHA2 = requests.post(echa_url, params=paramsECHA, data=dataECHA, headers=headersEDass, stream=True)
        if responseECHA2.status_code == 200:
            ECHA_database_bytes = BytesIO(responseECHA2.content)
            logging.info(f"Downloaded {echa_url}")
            responseECHA2.close()
            return ECHA_database_bytes
        else:
            logging.info(f"Failed to download {echa_url}. Status code:", responseECHA2.status_code)
            responseECHA2.close()
            return None
    except requests.exceptions.RequestException as e:
        logging.error(f"Network error while accessing {echa_url}: {e}")
        return None


# -------- Function to download a list from the ECHA-CHEM website url --------
async def download_echachem_list(list_url):
    """Click the 'Download full list' button and capture the file into BytesIO."""

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=['--disable-blink-features=AutomationControlled']
        )
        context = await browser.new_context(
            user_agent=random.choice(user_agents_list),
            viewport={'width': 1280, 'height': 720},
            device_scale_factor=1
        )
        page = await context.new_page()

        await page.goto(list_url, wait_until="networkidle")

        # Handle cookie/consent banner if present
        try:
            consent_button = page.locator('button:has-text("I accept the terms")')
            if await consent_button.is_visible():
                await consent_button.click()
                await page.wait_for_load_state("networkidle")
        except Exception as e:
            print(f"[WARN] Consent button not found: {e}")

        # Wait for the download button to appear before attempting click
        download_btn = page.locator('button:has-text("Download full list")')
        try:
            await download_btn.wait_for(state="visible", timeout=15000)
        except Exception:
            await browser.close()
            raise RuntimeError(f"'Download full list' button not found on page: {list_url}")

        # Intercept the download and click the button simultaneously
        try:
            async with page.expect_download(timeout=30000) as download_info:
                await download_btn.click()
            download = await download_info.value
        except Exception as e:
            await browser.close()
            raise RuntimeError(f"Download did not start after clicking the button: {e}")

        # Read the downloaded file directly into BytesIO (never touches disk)
        try:
            stream = await download.path()  # temp path Playwright wrote it to
            if stream is None:
                raise FileNotFoundError("Download path is None — the file may have failed to download.")
            echachem_bytes = BytesIO()
            with open(stream, "rb") as f:
                echachem_bytes.write(f.read())
            echachem_bytes.seek(0)
        except Exception as e:
            await browser.close()
            raise RuntimeError(f"Failed to read downloaded file into BytesIO: {e}")

        await browser.close()

    print(f"[✓] Downloaded '{download.suggested_filename}' into BytesIO.")
    return echachem_bytes


def process_data(file):
    logging.info("Started ED screener process")
    # Load file with CAS (or other input strings)
    CASallpd = pd.read_excel(file, engine="openpyxl")

    # Extract raw values (drop NaNs) preserving original order
    raw_values = CASallpd[selected_col].dropna().tolist()

    chem_ids = []   # combined valid identifiers (CAS + EC), unique in original order
    valid_cas = []
    valid_ec = []
    invalid_ids = []
    seen = set()
    order = []  # Save the order for reordering later on (needs to be unique values)

    # Quality check on IDs (eg CAS, EC)
    for raw in raw_values:
        norm = normalize_id(raw)
        if not norm:
            continue  # skip truly empty after normalization
        if norm in seen:
            continue  # keep original order, skip duplicates

        order.append(norm)
        if is_valid_cas(norm):
            chem_ids.append(norm)
            valid_cas.append(norm)
            seen.add(norm)
        elif is_valid_ec(norm):
            chem_ids.append(norm)
            valid_ec.append(norm)
            seen.add(norm)
        else:
            invalid_ids.append(norm)
    N_IDs = len(chem_ids)
    logging.info(f"Total input rows (non-NaN): {len(raw_values)}")
    logging.info(f"Valid CAS: {len(valid_cas)} | Valid EC: {len(valid_ec)} | Invalid: {len(invalid_ids)}")
    logging.info(f"Unique valid identifiers kept (order preserved): {len(chem_ids)}")
    st.write(f'The following ID numbers were invalid and therefore not screened: {invalid_ids}')

    # Create dictionary to save screening output in
    clp_info = [{"id": i + 1} for i in range(N_IDs)]  # Create list of dictionaries with length number of valid CAS numbers
    now = datetime.now()
    for i, entry in enumerate(clp_info):  # Add CAS and date to all entries
        entry["Input"] = chem_ids[i]
        entry["Date collected"] = now.strftime("%d/%m/%Y %H:%M:%S")
    # List of names to add as keys
    key_names = [
        "CAS", "EC", "Name ECHA-CHEM", "ECHA-CHEM checked", "REACH tonnage band", "On C&L?", "Entries C&L",
        "C&L URL", "C&L Type", "Joint Entries", "Classification - Hazard classes",
        "Classification - Hazard statements", "Classification - Organs/ExposureRoute",
        "Labeling - Hazard statements", "Labeling - Supplementary Hazard statements",
        "Labeling - Organs/ExposureRoute", "Specific concentration limits", "M-factors", "C&L notes",
        "ED PPP: Yes/No", "ED PPP: Status", "ED PPP: Conclusion HH", "ED PPP: Conclusion non-TO",
        "ED PPP: EFSA conclusion link",
        "BPR: Yes/No", "BPR: ED HH", "BPR: ED ENV",
        "ED Assessment List: Yes/No", "ED Assessment List: Outcome",
        "ED Assessment List: Status", "ED Assessment List: Authority", "ED Assessment List: Last updated",
        "SVHC: Yes/No", "SVHC: Reason", "SVHC: Date Inclusion", "SVHC: Decision",
        "Food additive: Yes/No", "Food additive: E number", "Food flavourings: Yes/No", "Food flavourings: FL",
        "SVHC intent: Yes/No", "SVHC intent: Status", "SVHC intent: Scope", "SVHC intent: Last updated",
        "PACT: Yes/No", "PACT: SEv", "PACT: SEv link", "PACT: DEv", "PACT: DEv link", "PACT: ED", "PACT: ED link",
        "PACT: ARN", "PACT: ARN link", "PACT: PBT", "PACT: PBT link", "PACT: CLH", "PACT: CLH link", "PACT: SVHC",
        "PACT: SVHC link",
        "CoRAP: Yes/No", "CoRAP: Initial grounds of Concern", "CoRAP: Status", "CoRAP: Latest update"
    ]
    # Add empty key-value pairs using dictionary unpacking
    clp_info = [{**entry, **{key: "-" for key in key_names}} for entry in clp_info]

    # Set up for requests (webscraping)
    user_agents_list = [
        'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.83 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:91.0) Gecko/20100101 Firefox/91.0',
        'Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15A372 Safari/604.1',
        'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_2_3) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15',
        'Mozilla/5.0 (Linux; Android 10; SM-G973F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.78 Mobile Safari/537.36',
        'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:98.0) Gecko/20100101 Firefox/98.0',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/117.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 13_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Linux; Android 11; Pixel 5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.131 Mobile Safari/537.36',
        'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
        'Mozilla/5.0 (Linux; Android 12; SM-A525F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.5845.92 Mobile Safari/537.36',
        'Mozilla/5.0 (iPad; CPU OS 15_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.5 Mobile/15E148 Safari/604.1',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 12_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15',
        'Mozilla/5.0 (Windows NT 10.0; ARM64; rv:110.0) Gecko/20100101 Firefox/110.0'
    ]
    # Function to download lists from the ECHA website

    #### LOAD DATA SOURCES ####
    logging.info("Loading databases")
    st.info("Loading databases")

    # >>> FROM EFSA WEBSITE <<<#
    # PPP ED
    efsaPPP_url = "https://www.efsa.europa.eu/en/applications/pesticides"
    PPP_ED_string = "overview-endocrine-disrupting-assessment-pesticide-active-substances"
    responseEFSA = requests.get(efsaPPP_url, headers={'User-Agent': random.choice(user_agents_list)})
    ED_PPP = None
    PPP_database_bytes = None
    if responseEFSA.status_code == 200:
        soupEFSA = BeautifulSoup(responseEFSA.text, "html.parser")
        matching_links = [link.get("href") for link in soupEFSA.find_all("a", href=True)
                          if PPP_ED_string in link.get("href") and link.get("href").endswith(('.xls', '.xlsx'))]
        if matching_links:
            file_url = requests.compat.urljoin(efsaPPP_url, matching_links[0])
            ED_PPP = requests.get(file_url)
            if ED_PPP.status_code == 200:
                PPP_database_bytes = BytesIO(ED_PPP.content)
                logging.info("Downloaded EFSA PPP ED database")
    workbook = openpyxl.load_workbook(PPP_database_bytes)
    sheetPPP = workbook.worksheets[0]
    st.write("PPP xlsx loaded")

    # >>> FROM ECHA-CHEM <<<
    # ED assessment list
    EDass_url = "https://chem.echa.europa.eu/activity-lists/edAssessment"
    EDass_database_bytes = None
    EDass_database_bytes = asyncio.get_event_loop().run_until_complete(download_echachem_list(EDass_url))
    workbookEDass = openpyxl.load_workbook(EDass_database_bytes, data_only=True)
    first_sheetEDass = workbookEDass.worksheets[0]
    st.write("EDass xlsx loaded")
    # SVHC list = Candidate list
    SVHC_url = "https://chem.echa.europa.eu/obligation-lists/candidateList"
    SVHC_database_bytes = None
    SVHC_database_bytes = asyncio.get_event_loop().run_until_complete(download_echachem_list(SVHC_url))
    workbookSVHC = openpyxl.load_workbook(SVHC_database_bytes, data_only=True)
    first_sheetSVHC = workbookSVHC.worksheets[0]
    st.write("SVHC xlsx loaded")
    # SVHC intent database
    SVHCintent_url = "https://chem.echa.europa.eu/activity-lists/svhcIdentification"
    SVHC_database_bytes = None
    SVHC_database_bytes = asyncio.get_event_loop().run_until_complete(download_echachem_list(SVHCintent_url))
    workbookSVHC = openpyxl.load_workbook(SVHC_database_bytes, data_only=True)
    first_sheetSVHC = workbookSVHC.worksheets[0]
    st.write("SVHC intent xlsx loaded")
    # CoRAP database incl Substance Evaluation (SEv, previously on PACT)
    CoRAP_url = "https://chem.echa.europa.eu/activity-lists/substanceEvaluation"
    CoRAP_database_bytes = None
    CoRAP_database_bytes = asyncio.get_event_loop().run_until_complete(download_echachem_list(CoRAP_url))
    workbookCoRAP = openpyxl.load_workbook(CoRAP_database_bytes, data_only=True)
    first_sheetCoRAP = workbookCoRAP.worksheets[0]
    st.write("CoRAP xlsx loaded")
    # Dossier Evaluation list (DEv, previously on PACT)
    DEv_url = "https://chem.echa.europa.eu/activity-lists/dossierEvaluation"
    DEv_database_bytes = None
    DEv_database_bytes = asyncio.get_event_loop().run_until_complete(download_echachem_list(DEv_url))
    workbookDEv = openpyxl.load_workbook(DEv_database_bytes, data_only=True)
    first_sheetDEv = workbookDEv.worksheets[0]
    st.write("DEv xlsx loaded")

    # Stil from old ECHA website (for now?)
    # Assessment of Regulatory Needs
    ARN_url = "https://www.echa.europa.eu/web/guest/assessment-regulatory-needs"
    ARN_database_bytes = download_echa_list(ARN_url, user_agents_list, source="ARN")
    workbookARN = openpyxl.load_workbook(ARN_database_bytes, data_only=True)
    first_sheetARN = workbookARN.worksheets[0]
    st.write("ARN xlsx loaded")


    # >>> FROM UPLOADED DATABASES <<<
    # BPR ED
    if file_BPR_ED is not None:
        workbookBPR = openpyxl.load_workbook(file_BPR_ED)
    else:
        st.warning("Please upload an Excel file for BPR ED.")
    logging.info(f"BPR ED list loaded successfully")
    # Food additives
    if file_food_add is not None:
        workbook_food_add = openpyxl.load_workbook(file_food_add)
    else:
        st.warning("Please upload an Excel file for food additives.")
    logging.info(f"Food additives list loaded successfully")
    # Food flavourings
    if file_food_flav is not None:
        workbook_food_flav = openpyxl.load_workbook(file_food_flav)
    else:
        st.warning("Please upload an Excel file for food flavourings.")
    logging.info(f"Food flavourings list loaded successfully")

    # C&L info from NextSDS API (Using JOBS)
    # Load API key
    if api_file is not None:
        api_key = api_file.read().decode("utf-8").strip()

    def chunk_list(lst, chunk_size):
        for i in range(0, len(lst), chunk_size):
            yield lst[i:i + chunk_size]

    logging.info("Starting nextSDS API")
    st.write("Checking ECHA-CHEM API")

    start_url = "https://api.nextsds.com/jobs/start"
    status_url = "https://api.nextsds.com/jobs/retrieve"
    headers = {
        "accept": "application/json",
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }

    # Step 1: Submit all jobs
    jobs = []
    for idx, cas_chunk in enumerate(chunk_list(chem_ids, 100)):
        data = {
            "taskId": "echa-api",
            "payload": cas_chunk
        }
        try:
            response = requests.post(start_url, headers=headers, json=data)
            if response.status_code == 200:
                job_id = response.json().get("id")
                jobs.append({"id": job_id, "index": idx + 1, "done": False, "output": None})
                st.write(f"Chunk {idx + 1}: Job submitted successfully: {job_id}")
                logging.info(f"Chunk {idx + 1}: Job submitted successfully: {job_id}")
            else:
                st.write(f"Chunk {idx + 1}: Failed to submit job")
                logging.info(f"Chunk {idx + 1}: Failed to submit job")
        except Exception as e:
            st.write(f"Chunk {idx + 1}: Exception during job submission: {str(e)}")
            logging.info(f"Chunk {idx + 1}: Exception during job submission: {str(e)}")

    # Step 2: Monitor all jobs in one loop
    while not all(job["done"] for job in jobs):
        time.sleep(5)
        for job in jobs:
            if job["done"]:
                continue
            try:
                status_response = requests.get(f"{status_url}/{job['id']}", headers=headers)
                if status_response.status_code == 200:
                    status_data = status_response.json()
                    job_status = status_data.get("status")
                    st.write(f"Chunk {job['index']}: Job status: {job_status}")
                    if job_status not in ["STARTED", "EXECUTING","DEQUEUED"]:
                        job["done"] = True
                        job["output"] = status_data.get("output", [])
                elif status_response.status_code in [400, 404]:
                    st.write(f"Chunk {job['index']}: Job error ({status_response.status_code})")
                    job["done"] = True
            except Exception as e:
                st.write(f"Chunk {job['index']}: Exception during status check: {str(e)}")

    # Step 3: Combine all outputs
    CnL_json = []
    for job in jobs:
        if job["output"]:
            CnL_json.extend(job["output"])

    logging.info("All jobs completed")
    st.write("All JSON chunks successfully retrieved and combined")

    #### LOOP OVER ALL CAS NUMBERS AND SCREEN SOURCES ####

    # Load all Excel files for using later
    try:
        workbook = openpyxl.load_workbook(PPP_database_bytes)
        sheetPPP = workbook.worksheets[0]
    except Exception as e:
        st.error(f"❌ Failed to load PPP Excel file: {e}")
        st.stop()
    try:
        workbookEDass = openpyxl.load_workbook(EDass_database_bytes)
        first_sheetEDass = workbookEDass.worksheets[0]
    except Exception as e:
        st.error(f"❌ Failed to load EDassessment Excel file: {e}")
        st.stop()
    try:
        workbookSVHC = openpyxl.load_workbook(SVHC_database_bytes)
        first_sheetSVHC = workbookSVHC.worksheets[0]
    except Exception as e:
        st.error(f"❌ Failed to load SVHC Excel file: {e}")
        st.stop()
    try:
        workbookSVHC_intent = openpyxl.load_workbook(SVHCintent_database_bytes)
        first_sheetSVHC_intent = workbookSVHC_intent.worksheets[0]
    except Exception as e:
        st.error(f"❌ Failed to load SVHCintent Excel file: {e}")
        st.stop()
    try:
        workbookPACT = openpyxl.load_workbook(PACT_database_bytes)
        first_sheetPACT = workbookPACT.worksheets[0]
    except Exception as e:
        st.error(f"❌ Failed to load PACT Excel file: {e}")
        st.stop()
    try:
        workbookCoRAP = openpyxl.load_workbook(CoRAP_database_bytes)
        first_sheetCoRAP = workbookCoRAP.worksheets[0]
    except Exception as e:
        st.error(f"❌ Failed to load CoRAP Excel file: {e}")
        st.stop()

    # st.write(CnL_json)
    i = 0
    while i < len(clp_info):
        st.write(f"Checking chemical: {clp_info[i]["Input"]}")

        #### ECHA-CHEM C&L from NEXTSDS-API ####
        # Find the entries matching the input string in the json
        matching_entries = [entryJSON for entryJSON in CnL_json if
                            entryJSON.get("casNumber") == clp_info[i]["Input"] or entryJSON.get("identifier") == clp_info[i]["Input"]]
        if len(matching_entries) > 1:  # If there are multiple hits for input string (eg multiple EC for CAS)
            if clp_info[i]["ECHA-CHEM checked"] == "-":  # If this is the first time the input string is checked
                clp_info[i]["ECHA-CHEM checked"] = 0  # Indicate that this is the first of multiple hits
                for n, entry in enumerate(matching_entries[1:], start=1):  # Skip the first match
                    copied_entry = copy.deepcopy(clp_info[i])  # Deep copy the original entry
                    copied_entry.update(entry)  # Merge with the new matching entry
                    copied_entry["ECHA-CHEM checked"] = n
                    clp_info.insert(i + n, copied_entry)  # Insert at position i+n
            entry = matching_entries[clp_info[i]["ECHA-CHEM checked"]]  # Use the corresponding entry in the json
        else:
            entry = matching_entries[0]  # Standard, take the first entry if only one hit
        # Extract required info from response json
        try:
            if entry.get("found") == False or len(entry.get("industryClassification"))==0:  # If the chemical was NOT found on C&L or if there are no industry classifications
                clp_info[i]["On C&L?"] = "No"
                # Add "not found" explicity for a few columns (easier for further use of output)
                for classname in ["Classification - Hazard classes","Classification - Hazard statements","Labeling - Hazard statements"]:
                    clp_info[i][classname] = "Not found on C&L"
            else:  # If the chemical was found on C&L (then there is no "found" entry)
                clp_info[i]["On C&L?"] = "Yes"
                clp_info[i]["CAS"] = entry.get("cas")
                clp_info[i]["EC"] = entry.get("ecNumber")
                clp_info[i]["Name ECHA-CHEM"] = entry.get("name")
                clp_info[i]["REACH tonnage band"] = entry.get("tonnageBand")
                # If harmonised classification, give this
                if entry.get("type") == "harmonised":
                    clp_info[i]["C&L Type"] = "Harmonised C&L"
                    clp_info[i]["C&L URL"] = "https://chem.echa.europa.eu/" + entry.get("rmlId") + "/harmonised"
                    clp_info[i]["Entries C&L"] = entry.get("totalIndustryClassifications")
                else:  # Self-classification by industry
                    clp_info[i]["C&L Type"] = "Notified C&L"
                    clp_info[i]["C&L URL"] = "https://chem.echa.europa.eu/" + entry.get("rmlId") + "/self-classified"
                    clp_info[i]["Entries C&L"] = entry.get("totalIndustryClassifications")
                    clp_info[i]["Joint Entries"] = entry.get("industryClassification")[0]["dataSource"]
                clp_info[i]["Classification - Hazard classes"] = entry.get("hazards")["hazardClasses"]
                clp_info[i]["Classification - Hazard statements"] = entry.get("hazards")["statements"]
                clp_info[i]["Classification - Organs/ExposureRoute"] = entry.get("hazards")["targetOrgsAndRoutes"]
                # Labelling
                labelling = entry.get("labelling", [])
                hazard_codes = [
                    item["hazardStatement"]["hazardStatementCode"]
                    for item in labelling
                    if "hazardStatement" in item and "hazardStatementCode" in item["hazardStatement"]
                ]
                clp_info[i]["Labeling - Hazard statements"] = ", ".join(hazard_codes)
                # clp_info[i]["Labeling - Hazard statements"] = entry.get("hazards")["statements"]
                # clp_info[i]["Labeling - Supplementary Hazard statements"] = entry.get("labelling")["targetOrgsAndRoutes"]
                # clp_info[i]["Labeling - Organs/ExposureRoute"] = entry.get("labelling")["targetOrgsAndRoutes"]
                clp_info[i]["Specific concentration limits"] = entry.get("hazards")["scl"]
                mfactor = entry.get("mFactor", {})
                if mfactor:
                    items = mfactor.get("items", [])
                    if isinstance(items, list) and items:
                        mfactor_strings = []
                        for item in items:
                            acute = item.get("mfactorAcute", "-")
                            chronic = item.get("mfactorChronic", "-")
                            mfactor_strings.append(f"Acute: {acute}; Chronic: {chronic}")
                        clp_info[i]["M-factors"] = " | ".join(mfactor_strings)
                    elif isinstance(mfactor, dict):
                        clp_info[i]["M-factors"] = "Acute: " + str(mfactor.get("mfactorAcute", "-")) + \
                                                   "; Chronic: " + str(mfactor.get("mfactorChronic", "-"))
                if entry.get("notes"):
                    clp_info[i]["C&L notes"] = entry.get("notes")[0]["note"]["noteCode"] + ": " + \
                                               entry.get("notes")[0]["note"]["noteText"]
            logging.info("Finished Next-SDS API")
        except:
            st.write(matching_entries)

        # Check PPP ED list
        if PPP_database_bytes:
            found_PPP_ED = False
            for row in sheetPPP.iter_rows(min_row=1, max_row=sheetPPP.max_row):
                for cell in row:
                    val = str(cell.value).strip()
                    if val != "-" and val in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):
                        clp_info[i]["ED PPP: Yes/No"] = "Yes"
                        clp_info[i]["ED PPP: Status"] = sheetPPP[f"H{cell.row}"].value
                        clp_info[i]["ED PPP: Conclusion HH"] = sheetPPP[f"I{cell.row}"].value
                        clp_info[i]["ED PPP: Conclusion non-TO"] = sheetPPP[f"J{cell.row}"].value
                        clp_info[i]["ED PPP: EFSA conclusion link"] = sheetPPP[f"N{cell.row}"].value
                        found_PPP_ED = True
                        break
                if found_PPP_ED:
                    break
            if not found_PPP_ED:
                clp_info[i]["ED PPP: Yes/No"] = "No"
        else:
            logging.info("No PPP database")

        # Check ECHA ED Assessment list
        if EDass_database_bytes:
            # Search for a specific string in the first sheet
            found_valueEDass = None
            for rowExcel_EDass in first_sheetEDass.iter_rows(min_row=1, max_row=first_sheetEDass.max_row, values_only=False):
                for cell in rowExcel_EDass:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        EDass_authority = first_sheetEDass[f"E{cell.row}"].value
                        found_valueEDass = EDass_authority
                        EDass_status = first_sheetEDass[f"F{cell.row}"].value
                        EDass_outcome = first_sheetEDass[f"G{cell.row}"].value
                        EDass_lastupdate = first_sheetEDass[f"H{cell.row}"].value
                        clp_info[i]["ED Assessment List: Yes/No"] = "Yes"
                        clp_info[i]["ED Assessment List: Outcome"] = EDass_outcome
                        clp_info[i]["ED Assessment List: Status"] = EDass_status
                        clp_info[i]["ED Assessment List: Authority"] = EDass_authority
                        clp_info[i]["ED Assessment List: Last updated"] = EDass_lastupdate
                        break
                if found_valueEDass:
                    break
            if not found_valueEDass:
                clp_info[i]["ED Assessment List: Yes/No"] = "No"
        else:
            logging.info("No ECHA ED database")

        # Check SVHC list
        if SVHC_database_bytes:
            # Search for a specific string in the first sheet
            found_valueSVHC = None
            for rowExcel_SVHC in first_sheetSVHC.iter_rows(min_row=1, max_row=first_sheetSVHC.max_row, values_only=False):
                for cell in rowExcel_SVHC:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        SVHCreason = first_sheetSVHC[f"E{cell.row}"].value
                        found_valueSVHC = SVHCreason
                        SVHCdate = first_sheetSVHC[f"I{cell.row}"].value
                        SVHCdecision = first_sheetSVHC[f"J{cell.row}"].value
                        clp_info[i]["SVHC: Yes/No"] = "Yes"
                        clp_info[i]["SVHC: Reason"] = SVHCreason
                        clp_info[i]["SVHC: Date Inclusion"] = SVHCdate
                        clp_info[i]["SVHC: Decision"] = SVHCdecision
                        break
                if found_valueSVHC:
                    break
            if not found_valueSVHC:
                clp_info[i]["SVHC: Yes/No"] = "No"
        else:
            logging.info("No SVHC database")

        # Check SVHC intent list
        if SVHCintent_database_bytes:
            # Search for a specific string in the first sheet
            found_valueSVHC_intent = None
            for rowExcel_SVHC_intent in first_sheetSVHC_intent.iter_rows(min_row=1, max_row=first_sheetSVHC_intent.max_row, values_only=False):
                for cell in rowExcel_SVHC_intent:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        # Extract the value from column G of the same row
                        SVHC_intent_status = first_sheetSVHC_intent[f"G{cell.row}"].value
                        found_valueSVHC_intent = SVHC_intent_status
                        SVHC_intent_scope = first_sheetSVHC_intent[f"N{cell.row}"].value
                        SVHC_intent_lastupdated = first_sheetSVHC_intent[f"AA{cell.row}"].value
                        clp_info[i]["SVHC intent: Yes/No"] = "Yes"
                        clp_info[i]["SVHC intent: Status"] = SVHC_intent_status
                        clp_info[i]["SVHC intent: Scope"] = SVHC_intent_scope
                        clp_info[i]["SVHC intent: Last updated"] = SVHC_intent_lastupdated
                        break
                if found_valueSVHC_intent:
                    break
            if not found_valueSVHC_intent:
                clp_info[i]["SVHC intent: Yes/No"] = "No"
        else:
            logging.info("No SVHC intent database")

        # Check PACT list
        if PACT_database_bytes:
            # Search for a specific string in the first sheet
            found_valuePACT = None
            for rowExcel_PACT in first_sheetPACT.iter_rows(min_row=1, max_row=first_sheetPACT.max_row, values_only=False):
                for cell in rowExcel_PACT:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        # print(f"Found '{CAS}' in row {cell.row}, column {cell.column} on PACT list.")
                        # Extract the SEv from column E of the same row
                        PACT_SEv = first_sheetPACT[f"E{cell.row}"].value
                        found_valuePACT = PACT_SEv
                        PACT_SEv_link = first_sheetPACT[f"F{cell.row}"].value
                        PACT_DEv = first_sheetPACT[f"I{cell.row}"].value
                        PACT_DEv_link = first_sheetPACT[f"J{cell.row}"].value
                        PACT_ED = first_sheetPACT[f"K{cell.row}"].value
                        PACT_ED_link = first_sheetPACT[f"L{cell.row}"].value
                        PACT_ARN = first_sheetPACT[f"M{cell.row}"].value
                        PACT_ARN_link = first_sheetPACT[f"N{cell.row}"].value
                        PACT_PBT = first_sheetPACT[f"G{cell.row}"].value
                        PACT_PBT_link = first_sheetPACT[f"H{cell.row}"].value
                        PACT_CLH = first_sheetPACT[f"O{cell.row}"].value
                        PACT_CLH_link = first_sheetPACT[f"P{cell.row}"].value
                        PACT_SVHC = first_sheetPACT[f"Q{cell.row}"].value
                        PACT_SVHC_link = first_sheetPACT[f"R{cell.row}"].value
                        clp_info[i]["PACT: Yes/No"] = "Yes"
                        clp_info[i]["PACT: SEv"] = PACT_SEv
                        clp_info[i]["PACT: SEv link"] = PACT_SEv_link
                        clp_info[i]["PACT: DEv"] = PACT_DEv
                        clp_info[i]["PACT: DEv link"] = PACT_DEv_link
                        clp_info[i]["PACT: ED"] = PACT_ED
                        clp_info[i]["PACT: ED link"] = PACT_ED_link
                        clp_info[i]["PACT: ARN"] = PACT_ARN
                        clp_info[i]["PACT: ARN link"] = PACT_ARN_link
                        clp_info[i]["PACT: PBT"] = PACT_PBT
                        clp_info[i]["PACT: PBT link"] = PACT_PBT_link
                        clp_info[i]["PACT: CLH"] = PACT_CLH
                        clp_info[i]["PACT: CLH link"] = PACT_CLH_link
                        clp_info[i]["PACT: SVHC"] = PACT_SVHC
                        clp_info[i]["PACT: SVHC link"] = PACT_SVHC_link
                        break
                if found_valuePACT:
                    break
            if not found_valuePACT:
                clp_info[i]["PACT: Yes/No"] = "No"
        else:
            logging.info("No PACT database")

        # Check CoRAP
        if CoRAP_database_bytes:
            # Search for a specific string in the first sheet
            found_valueCoRAP = None
            for rowExcel_CoRAP in first_sheetCoRAP.iter_rows(min_row=1, max_row=first_sheetCoRAP.max_row, values_only=False):
                for cell in rowExcel_CoRAP:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]): # Check both CAS or EC
                        CoRAPgrounds = first_sheetCoRAP[f"H{cell.row}"].value
                        found_valueCoRAP = CoRAPgrounds
                        CoRAPstatus = first_sheetCoRAP[f"I{cell.row}"].value
                        CoRAPlastupdate = first_sheetCoRAP[f"J{cell.row}"].value
                        clp_info[i]["CoRAP: Yes/No"] = "Yes"
                        clp_info[i]["CoRAP: Initial grounds of Concern"] = CoRAPgrounds
                        clp_info[i]["CoRAP: Status"] = CoRAPstatus
                        clp_info[i]["CoRAP: Latest update"] = CoRAPlastupdate
                        break
                if found_valueCoRAP:
                    break
            if not found_valueCoRAP:
                clp_info[i]["CoRAP: Yes/No"] = "No"
        else:
            logging.info("No CoRAP database")

        # Check BPR ED list
        if workbookBPR:
            sheetBPR = workbookBPR["List of active substances"]
            # Search for a specific string in the first sheet
            found_valueBPR = None
            for rowExcel_BPR in sheetBPR.iter_rows(min_row=1, max_row=sheetBPR.max_row, values_only=False):
                for cell in rowExcel_BPR:
                    cell_value = str(cell.value).strip().replace("\u00A0", "").replace(" ","")
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        found_valueBPR = sheetBPR[f"K{cell.row}"].value
                        clp_info[i]["BPR: Yes/No"] = "Yes"
                        clp_info[i]["BPR: ED HH"] = sheetBPR[f"K{cell.row}"].value
                        clp_info[i]["BPR: ED ENV"] = sheetBPR[f"L{cell.row}"].value
                        break
                if found_valueBPR:
                    break
            if not found_valueBPR:
                clp_info[i]["BPR: Yes/No"] = "No"
        else:
            logging.info("No BPR ED database")

        # Check Food additives
        if workbook_food_add:
            sheet_food_add = workbook_food_add["List for EDscreener"]
            found_value_food_add = None
            for rowExcel_food_add in sheet_food_add.iter_rows(min_row=1, max_row=sheet_food_add.max_row, values_only=False):
                for cell in rowExcel_food_add:
                    cell_value = str(cell.value).strip()
                    # Based on CAS in column B
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        food_add_Enumber = sheet_food_add[f"B{cell.row}"].value
                        found_value_food_add = food_add_Enumber
                        clp_info[i]["Food additive: Yes/No"] = "Yes"
                        clp_info[i]["Food additive: E number"] = food_add_Enumber
                        break
                    # Also check based on EC in column A (and check if EC is existing)
                    ECnr = clp_info[i]["EC"]
                    if cell.value and ECnr in str(cell.value).strip() and len(ECnr) > 1:
                        food_add_Enumber = sheet_food_add[f"A{cell.row}"].value
                        found_value_food_add = food_add_Enumber
                        clp_info[i]["Food additive: Yes/No"] = "Yes"
                        clp_info[i]["Food additive: E number"] = food_add_Enumber
                        break
                if found_value_food_add:
                    break
            if not found_value_food_add:
                clp_info[i]["Food additive: Yes/No"] = "No"
        else:
            logging.info("No Food additives database")

        #  Food flavourings
        if workbook_food_flav:
            sheet_food_flav = workbook_food_flav["List for EDscreener"]
            # Search for a specific string in the sheet
            found_value_food_flav = None
            for rowExcel_food_flav in sheet_food_flav.iter_rows(min_row=1, max_row=sheet_food_flav.max_row, values_only=False):
                for cell in rowExcel_food_flav:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        found_value_food_flav = "Yes"
                        food_flav_FL = sheet_food_flav[f"A{cell.row}"].value
                        clp_info[i]["Food flavourings: Yes/No"] = "Yes"
                        clp_info[i]["Food flavourings: FL"] = "FL " + food_flav_FL
                        break
                if found_value_food_flav:
                    break
            if not found_value_food_flav:
                clp_info[i]["Food flavourings: Yes/No"] = "No"
        else:
            logging.info("No Food flavourings database")

        # Finalize the loop per chemical
        logging.info(f"Processed {i+1}/{len(clp_info)}: {clp_info[i]["CAS"]}")
        st.write(f"Processed {i+1}/{len(clp_info)}: {clp_info[i]["CAS"]}")

        i += 1  # Update in the while loop

    # Make panda dataframe for saving to Excel
    df = pd.DataFrame(clp_info)
    # Reorder to the original list order
    df["Input"] = pd.Categorical(df["Input"], categories=order, ordered=True)
    df = df.sort_values(by="Input").reset_index(drop=True)

    output_excel = BytesIO()
    df.to_excel(output_excel, index=False, engine="openpyxl")
    output_excel.seek(0)


    ### Add Summary sheet to Excel output ####
    # Load workbook
    wb = load_workbook(output_excel)

    # Make the URLs clickable, Loop from the second row to the last row
    ws = wb["Sheet1"]
    columns = [11, 50, 52, 56, 60]  # C&L, PACT SEv, DEv, ARN, CLH
    link_style = Font(color="0000FF", underline="single")
    for row in range(2, ws.max_row + 1):
        for col in columns:
            cell = ws.cell(row=row, column=col)
            cell.hyperlink = cell.value
            cell.font = link_style

    # Check if the "Summary" sheet exists; if not, create it
    if "Summary" not in wb.sheetnames:
        ws = wb.create_sheet(title="Summary")
    else:
        ws = wb["Summary"]

    # Headers to be added in the "Summary" sheet
    headers = [
        ["", "", "", "", "", "Evaluated for ED in", "", "", "", "", "", "Also found in", "", ""],
        ["Name (ECHA-CHEM)", "Input", "CAS number", "EC number", "Classification", "ED assessment",
         "On BPR/PPPR list (for ED-HH; for ED-ENV)", "REACH SVHC candidate", "REACH SVHC intent",
         "CORAP List", "PACT: DEv", "PACT: ARN", "Food lists", "Summary Harmonized", "Summary self-classified"]
    ]
    # Arrange headers and formatting
    for row_index, row_data in enumerate(headers, start=1):
        for col_index, cell_data in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=cell_data)
            # Make the font bold for the header cells
            cell.font = Font(bold=True)
    ws.merge_cells("F1:J1")

    # Insert formulas for summary & Autofill down
    last_row = wb["Sheet1"].max_row
    for row in range(3, last_row + 2):
        ws[f"A{row}"] = f"=Sheet1!B{row - 1}"  # Input
        ws[f"B{row}"] = f"=Sheet1!F{row - 1}"  # Name
        ws[f"C{row}"] = f"=Sheet1!D{row - 1}"  # CAS
        ws[f"D{row}"] = f"=Sheet1!E{row - 1}"  # EC
        ws[f"E{row}"] = f'=IF(Sheet1!N{row - 1}<>"",Sheet1!N{row - 1},"Not classified")'  # Classification
        # ED assessment list
        ws[
            f"F{row}"] = f'=CONCATENATE(Sheet1!AE{row - 1},IF(OR(Sheet1!AE{row - 1}="No",Sheet1!AE{row - 1}=""),""," ("&Sheet1!AF{row - 1}&")"))'
        # BPR/PPP ED?
        ws[
            f"G{row}"] = f'="BPR: "&IF(Sheet1!AB{row - 1}="Yes","Yes (HH: " &Sheet1!AC{row - 1}& "; ENV: " &Sheet1!AD{row - 1}& ")","No")&' \
                         f'"; PPR: "&IF(Sheet1!W{row - 1}="Yes","Yes (HH: "&Sheet1!Y{row - 1} &"; ENV: " &Sheet1!Z{row - 1}& ")","No")'
        # REACH SVHC candidate ED?
        ws[f"H{row}"] = f'=IF(Sheet1!AJ{row - 1}="Yes","Yes: " & Sheet1!AK{row - 1},"No")'
        # REACH SVHC intent ED?
        ws[f"I{row}"] = f'=IF(Sheet1!AR{row - 1}="Yes","Yes: " & Sheet1!AS{row - 1},"No")'
        # CORAP list ED?
        ws[f"J{row}"] = f'=Sheet1!BK{row - 1}&" ("&Sheet1!BL{row - 1}&"; "&Sheet1!BM{row - 1}&")"'
        # PACT: Dev
        ws[f"K{row}"] = f'=Sheet1!AY{row - 1}'
        # PACT: ARN
        ws[f"L{row}"] = f'=Sheet1!BC{row - 1}'
        # Food additives/flavourings
        ws[
            f"M{row}"] = f'=IF(OR(Sheet1!AN{row - 1}="Yes",Sheet1!AP{row - 1}="Yes"),"Yes (" & Sheet1!AO{row - 1} & "; " & Sheet1!AQ{row - 1} & ")", "No")'

    # Determine classification interpretation before adding to Excel
    # Function to determine classification summary
    def determine_classification(value):
        if pd.isna(value) or value.strip() == '':
            return '-'
        matched_classifications = []
        for outcome, codes in classification_mapping.items():
            if any(code in value for code in codes):
                matched_classifications.append(outcome)
        return ', '.join(matched_classifications) if matched_classifications else 'other classification'

    # Set up mapping for classification
    classification_mapping = {
        'reproductive toxicity': ['H360', 'H360F', 'H360FD', 'H360Fd', 'H360Df', 'H361', 'H361f', 'H361d', 'H361fd','H362'],
        'STOT-RE': ['H372', 'H373'],
        'carcinogenicity': ['H350', 'H350i', 'H351'],
        'endocrine disruption':['EUH380', 'EUH381', 'EUH430', 'EUH431']
    }

    # Determine classification based on harmonized & self-classified column
    df['Harmonized C&L assessment'] = df['Classification - Hazard statements'].apply(determine_classification)
    df['Self-classified C&L assessment'] = df['Classification - Hazard statements'].apply(determine_classification)
    # If C&L Type is NOT 'Harmonized C&L', set 'Harmonized' column to '-'
    df.loc[df["C&L Type"] != "Harmonised C&L", "Harmonized C&L assessment"] = "-"
    # If C&L Type IS 'Harmonized C&L', set 'Self-class' column to '-'
    df.loc[df["C&L Type"] == "Harmonised C&L", "Self-classified C&L assessment"] = "-"

    # Add to Excel summary sheet
    for i, (val1, val2) in enumerate(zip(df['Harmonized C&L assessment'], df['Self-classified C&L assessment']),
                                     start=3):
        ws[f"N{i}"] = val1
        ws[f"O{i}"] = val2

    # Some formatting
    # Set column width to 12 for columns A to AZ
    for col in range(1, 106):  # 1 to 52 (A to BZ)
        col_letter = get_column_letter(col)
        # For sheet 1
        cell = wb["Sheet1"].cell(row=1, column=col)
        cell.alignment = Alignment(wrap_text=True)
        wb["Sheet1"].column_dimensions[col_letter].width = 12
        # For summary sheet
        for rowN in (1, 2):
            cell = ws.cell(row=rowN, column=col)
            cell.alignment = Alignment(wrap_text=True)
        ws.column_dimensions[col_letter].width = 15

    # Save final workbook
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    ### SAVING ####
    # Save to zip file
    st.session_state.log_stream.seek(0)
    log_bytes = BytesIO(st.session_state.log_stream.read().encode("utf-8"))
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        zip_file.writestr("EDscreener_results.xlsx", final_output.getvalue())
        zip_file.writestr("EDscreener_log.txt", log_bytes.getvalue())
        if PPP_database_bytes:
            zip_file.writestr("databases/EFSA_PPP_ED_Database.xlsx", PPP_database_bytes.getvalue())
        if EDass_database_bytes:
            zip_file.writestr("databases/ED assessment_Database.xlsx", EDass_database_bytes.getvalue())
        if SVHC_database_bytes:
            zip_file.writestr("databases/SVHC_Database.xlsx", SVHC_database_bytes.getvalue())
        if SVHCintent_database_bytes:
            zip_file.writestr("databases/SVHC intent_Database.xlsx", SVHCintent_database_bytes.getvalue())
        if PACT_database_bytes:
            zip_file.writestr("databases/PACT_Database.xlsx", PACT_database_bytes.getvalue())
        if CoRAP_database_bytes:
            zip_file.writestr("databases/CoRAP_Database.xlsx", CoRAP_database_bytes.getvalue())
        if workbookBPR:
            excel_buffer = BytesIO()    # Save workbook to a BytesIO buffer
            workbookBPR.save(excel_buffer)
            excel_buffer.seek(0)  # Reset buffer position to the beginning
            zip_file.writestr(f"databases/{file_BPR_ED.name}", excel_buffer.getvalue())
        if workbook_food_add:
            excel_buffer = BytesIO()    # Save workbook to a BytesIO buffer
            workbook_food_add.save(excel_buffer)
            excel_buffer.seek(0)  # Reset buffer position to the beginning
            zip_file.writestr(f"databases/{file_food_add.name}", excel_buffer.getvalue())
        if workbook_food_flav:
            excel_buffer = BytesIO()    # Save workbook to a BytesIO buffer
            workbook_food_flav.save(excel_buffer)
            excel_buffer.seek(0)  # Reset buffer position to the beginning
            zip_file.writestr(f"databases/{file_food_flav.name}", excel_buffer.getvalue())
        json_data = json.dumps(CnL_json, indent=2)
        json_bytes = json_data.encode('utf-8')
        if json_bytes:
            zip_file.writestr(f"databases/API json {datetime.now().strftime('%Y-%m-%d %H-%M')}.json", json_bytes)
    zip_buffer.seek(0)

    return zip_buffer

if uploaded_file:
    df_preview = pd.read_excel(uploaded_file, engine="openpyxl")
    st.write("### Preview of uploaded file")
    # Dropdown for column selection
    selected_col = st.selectbox(
        "Select the column containing CAS/EC numbers:",
        df_preview.columns
    )

    if st.button("Run Screener"):
        st.info("Processing started...")
        zip_result = process_data(uploaded_file)
        if zip_result:
            st.success("Processing finished!")
            st.download_button("Download All Results (ZIP)", zip_result, file_name=f"EDscreener_package_{datetime.now().strftime('%Y-%m-%d %H-%M')}.zip")

        else:
            # If processing failed, offer log download
            st.session_state.log_stream.seek(0)
            log_text = st.session_state.log_stream.read()
            st.download_button("Download Log Only", log_text, file_name="EDscreener_log.txt")
            st.error("Processing failed. You can download the log for details.")
