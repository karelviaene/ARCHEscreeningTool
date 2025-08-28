import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import pandas as pd
import os
import re
import time
import random
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from io import BytesIO
import openpyxl

class EDScreenerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ED Screener Tool")

        self.file_path = None
        self.folder_path = os.getcwd()

        # GUI Elements
        tk.Button(root, text="Select Excel File", command=self.select_file).pack()
        self.file_label = tk.Label(root, text="No file selected")
        self.file_label.pack()

        tk.Button(root, text="Select Output Folder", command=self.select_folder).pack()
        self.folder_label = tk.Label(root, text=f"Output folder: {self.folder_path}")
        self.folder_label.pack()

        tk.Button(root, text="Run Screener", command=self.run_screener).pack()
        self.status_label = tk.Label(root, text="")
        self.status_label.pack()

    def select_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.file_label.config(text=self.file_path)

    def select_folder(self):
        self.folder_path = filedialog.askdirectory()
        self.folder_label.config(text=f"Output folder: {self.folder_path}")

    def run_screener(self):
        if not self.file_path:
            messagebox.showerror("Error", "Please select an Excel file.")
            return
        self.status_label.config(text="Running...")
        threading.Thread(target=self.process_data).start()

    def process_data(self):
        try:
            CASallpd = pd.read_excel(self.file_path)
            if "CAS" not in CASallpd.columns:
                self.status_label.config(text="Error: 'CAS' column not found.")
                return

            CASall = CASallpd["CAS"].dropna().tolist()
            CASall = [re.sub(r'[^\d\-]', '', str(cas)) for cas in CASall]
            N_CAS = len(CASall)
            clp_info = [{"id": i+1, "Input": CASall[i]} for i in range(N_CAS)]

            # Add empty fields
            key_names = [
                 "Input","CAS","EC","Name ECHA-CHEM","ECHA-CHEM checked","REACH tonnage band","On C&L?","Entries C&L","C&L URL","C&L Type","Joint Entries","Classification - Hazard classes",
                 "Classification - Hazard statements","Classification - Organs/ExposureRoute",
                 "Labeling - Hazard statements", "Labeling - Supplementary Hazard statements",
                 "Labeling - Organs/ExposureRoute","Specific concentration limits","M-factors","C&L notes",
                 "ED PPP: Yes/No","ED PPP: Status","ED PPP: Conclusion HH","ED PPP: Conclusion non-TO",
                 "ED PPP: EFSA conclusion link",
                 "BPR: Yes/No","BPR: ED HH","BPR: ED ENV",
                 "ED Assessment List: Yes/No","ED Assessment List: Outcome",
                 "ED Assessment List: Status","ED Assessment List: Authority","ED Assessment List: Last updated",
                 "SVHC: Yes/No","SVHC: Reason","SVHC: Date Inclusion","SVHC: Decision",
                 "Food additive: Yes/No","Food additive: E number","Food flavourings: Yes/No","Food flavourings: FL",
                 "SVHC intent: Yes/No","SVHC intent: Status","SVHC intent: Scope","SVHC intent: Last updated",
                 "PACT: Yes/No","PACT: SEv","PACT: SEv link","PACT: DEv","PACT: DEv link","PACT: ED","PACT: ED link",
                 "PACT: ARN","PACT: ARN link","PACT: PBT","PACT: PBT link","PACT: CLH","PACT: CLH link","PACT: SVHC",
                 "PACT: SVHC link",
                 "CoRAP: Yes/No","CoRAP: Initial grounds of Concern","CoRAP: Status","CoRAP: Latest update"
                 ]
            for entry in clp_info:
                for key in key_names:
                    entry[key] = "-"

            # Create databases folder
            databases_folder = os.path.join(self.folder_path, "databases")
            os.makedirs(databases_folder, exist_ok=True)
            # Function to check if file is up to date (from today)
            def file_downloaded_today(path):
                if os.path.exists(path):
                    mod_time = datetime.fromtimestamp(os.path.getmtime(path)).date()
                    return mod_time == datetime.now().date()
                return False
            datetoday = datetime.now().strftime("%Y-%m-%d")

            # Download EFSA PPP ED Excel
            efsaPPP_url = "https://www.efsa.europa.eu/en/applications/pesticides"
            PPP_ED_string = "overview-endocrine-disrupting-assessment-pesticide-active-substances"
            responseEFSA = requests.get(efsaPPP_url)
            ED_PPP = None

            if responseEFSA.status_code == 200:
                soupEFSA = BeautifulSoup(responseEFSA.text, "html.parser")
                matching_links = [link.get("href") for link in soupEFSA.find_all("a", href=True)
                                  if PPP_ED_string in link.get("href") and link.get("href").endswith((".xls", ".xlsx"))]
                if matching_links:
                    file_url = requests.compat.urljoin(efsaPPP_url, matching_links[0])
                    ED_PPP = requests.get(file_url)
                    # Step 0: Check if file already was downloaded today
                    file_name = "PPP ED list " + datetoday + ".xlsx"
                    file_path_PPP_ED = os.path.join(databases_folder, file_name)
                    # Only download if file was not downloaded already today
                    if file_downloaded_today(file_path_PPP_ED):
                        self.status_label.config(text="PPP ED file already downloaded today.")
                    else:
                        if ED_PPP.status_code == 200:  # Download to databases folder
                            with open(file_path_PPP_ED, "wb") as file:
                                file.write(ED_PPP.content)
                            self.status_label.config(text=f"Saved PPP ED list: {file_path_PPP_ED}")
                        else:
                            self.status_label.config(text=f"Failed to download PPP ED file: {file_url}")

            # Process each CAS
            for i, entry in enumerate(clp_info):
                if ED_PPP and ED_PPP.status_code == 200:
                    excel_data = BytesIO(ED_PPP.content)
                    workbook = openpyxl.load_workbook(excel_data)
                    sheet = workbook.worksheets[0]
                    found = False
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                        for cell in row:
                            val = str(cell.value).strip()
                            if val in [entry["Input"]]:
                                entry["ED PPP: Yes/No"] = "Yes"
                                entry["ED PPP: Status"] = sheet[f"H{cell.row}"].value
                                entry["ED PPP: Conclusion HH"] = sheet[f"I{cell.row}"].value
                                entry["ED PPP: Conclusion non-TO"] = sheet[f"J{cell.row}"].value
                                entry["ED PPP: EFSA conclusion link"] = sheet[f"N{cell.row}"].value
                                found = True
                                break
                        if found:
                            break
                self.status_label.config(text=f"Processed {i+1}/{N_CAS}")

            # Save results
            df = pd.DataFrame(clp_info)
            now = datetime.now().strftime("%Y-%m-%d %H-%M")
            output_file = os.path.join(self.folder_path, f"EDscreener_export_{now}.xlsx")
            df.to_excel(output_file, index=False)
            self.status_label.config(text=f"Saved to {output_file}")

        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}")

# Run the app
if __name__ == "__main__":
    root = tk.Tk()
    app = EDScreenerApp(root)
    root.mainloop()