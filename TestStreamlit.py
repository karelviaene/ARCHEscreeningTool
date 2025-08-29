
import streamlit as st
import pandas as pd
import re
import requests
from bs4 import BeautifulSoup
from io import BytesIO, StringIO
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import logging
import zipfile
import random
import time
from datetime import datetime

st.title("ED Screener Tool")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])
file_BPR_ED = st.file_uploader("Upload BPR ED file (xlsx)", type=["xlsx"])
file_food_add = st.file_uploader("Upload Food additives Excel (xlsx)", type=["xlsx"])
file_food_flav = st.file_uploader("Upload Food flavourings Excel (xlsx)", type=["xlsx"])

# In-memory log stream
if "log_stream" not in st.session_state:
    st.session_state.log_stream = StringIO()
    logging.basicConfig(stream=st.session_state.log_stream, level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s')

def process_data(file):
    logging.info("Started ED screener process")
    CASallpd = pd.read_excel(file, engine="openpyxl")
    if "CAS" not in CASallpd.columns:
        st.error("Error: 'CAS' column not found.")
        return None

    CASall = CASallpd["CAS"].dropna().tolist()
    CASall = [re.sub(r'[^\d\-]', '', str(cas)) for cas in CASall]
    N_CAS = len(CASall)

    clp_info = [{"id": i+1, "Input": CASall[i]} for i in range(N_CAS)]
    key_names = [
        "Input", "CAS", "EC", "Name ECHA-CHEM", "ECHA-CHEM checked", "REACH tonnage band", "On C&L?", "Entries C&L",
        "C&L URL", "C&L Type", "Joint Entries", "Classification - Hazard classes",
        "Classification - Hazard statements", "Classification - Organs/ExposureRoute",
        "Labeling - Hazard statements", "Labeling - Supplementary Hazard statements",
        "Labeling - Organs/ExposureRoute", "Specific concentration limits", "M-factors", "C&L notes",
        "ED PPP: Yes/No", "ED PPP: Status", "ED PPP: Conclusion HH", "ED PPP: Conclusion non-TO",
        "ED PPP: EFSA conclusion link",
        "BPR: Yes/No", "BPR: ED HH", "BPR: ED ENV",
        "ED Assessment List: Yes/No", "ED Assessment List: Outcome",
        "ED Assessment List: Status", "ED Assessment List: Authority", "ED Assessment List: Last updated",
        "SVHC: Yes/No", "SVHC: Reason", "SVHC: Date Inclusion", "SVHC: Decision",
        "Food additive: Yes/No", "Food additive: E number", "Food flavourings: Yes/No", "Food flavourings: FL",
        "SVHC intent: Yes/No", "SVHC intent: Status", "SVHC intent: Scope", "SVHC intent: Last updated",
        "PACT: Yes/No", "PACT: SEv", "PACT: SEv link", "PACT: DEv", "PACT: DEv link", "PACT: ED", "PACT: ED link",
        "PACT: ARN", "PACT: ARN link", "PACT: PBT", "PACT: PBT link", "PACT: CLH", "PACT: CLH link", "PACT: SVHC",
        "PACT: SVHC link",
        "CoRAP: Yes/No", "CoRAP: Initial grounds of Concern", "CoRAP: Status", "CoRAP: Latest update"
    ]
    # Add empty key-value pairs using dictionary unpacking
    clp_info = [{**entry, **{key: "-" for key in key_names}} for entry in clp_info]

    # Set up for requests (webscraping)
    user_agents_list = [
        'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.83 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:91.0) Gecko/20100101 Firefox/91.0',
        'Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15A372 Safari/604.1',
        'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_2_3) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15',
        'Mozilla/5.0 (Linux; Android 10; SM-G973F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.78 Mobile Safari/537.36',
        'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:98.0) Gecko/20100101 Firefox/98.0',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/117.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 13_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Linux; Android 11; Pixel 5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.131 Mobile Safari/537.36',
        'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
        'Mozilla/5.0 (Linux; Android 12; SM-A525F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.5845.92 Mobile Safari/537.36',
        'Mozilla/5.0 (iPad; CPU OS 15_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.5 Mobile/15E148 Safari/604.1',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 12_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15',
        'Mozilla/5.0 (Windows NT 10.0; ARM64; rv:110.0) Gecko/20100101 Firefox/110.0'
    ]
    # Function to download lists from the ECHA website
    def download_echa_list(echa_url, user_agents_list,source):
        try:
            headers = {'User-Agent': random.choice(user_agents_list)}
            responseECHA = requests.get(echa_url, headers=headers)

            unique_substances = None
            if responseECHA.status_code == 200:
                soupECHA = BeautifulSoup(responseECHA.text, "html.parser")
                small_tag = soupECHA.find("small", class_="search-results")
                if small_tag:
                    text = small_tag.get_text(strip=True)
                    match = re.search(r"of\s+([\d,]+)\s+results", text)
                    if match:
                        if source == "PACT":
                            unique_substances = match.group(1).replace(",", "")
                        else:
                            unique_substances = match.group(1)
                soupECHA.decompose()
            responseECHA.close()
            if not unique_substances:
                logging.info("Could not determine the number of unique substances.")
                return None

            # Data or payload sent with the POST request+
            paramsECHA = {
                "p_p_id": "disslists_WAR_disslistsportlet",
                "p_p_lifecycle": "2",
                "p_p_state": "normal",
                "p_p_mode": "view",
                "p_p_resource_id": "exportResults",
                "p_p_cacheability": "cacheLevelPage"
            }
            dataECHA = {
                "_disslists_WAR_disslistsportlet_formDate": int(round(time.time() * 1000)),
                "_disslists_WAR_disslistsportlet_exportColumns": "name,ecNumber,casNumber,lec_submitter,prc_public_status,prc_conclusion,diss_update_date,dte_first_published",
                "_disslists_WAR_disslistsportlet_orderByCol": "diss_update_date",
                "_disslists_WAR_disslistsportlet_orderByType": "asc",
                "_disslists_WAR_disslistsportlet_searchFormColumns": "prc_public_status,prc_conclusion,lec_submitter,dte_intention,dte_assessment,diss_update_date",
                "_disslists_WAR_disslistsportlet_searchFormElements": "DROP_DOWN,DROP_DOWN,DROP_DOWN,DATE_PICKER,DATE_PICKER,DATE_PICKER",
                "_disslists_WAR_disslistsportlet_total": unique_substances,
                "_disslists_WAR_disslistsportlet_exportType": "xls"
            }
            headersEDass = {
                "User-Agent": random.choice(user_agents_list),
                "Content-Type": "application/x-www-form-urlencoded",
            }

            # Adjust based on source
            if source == "PACT":
                paramsECHA["p_p_id"] = "disslists_WAR_disslistsportlet"
                dataECHA = {
                    "_disspact_WAR_disspactportlet_formDate": int(round(time.time() * 1000)),
                    "_disspact_WAR_disspactportlet_exportColumns": "name,ecNumber,casNumber,DISLIST_CORAP,DISLIST_PBT,DISLIST_DOSSIER_EVALUATION,DISLIST_ED,DISLIST_ARN,DISLIST_ROI_CLH,DISLIST_ROI_SVHC,DISLIST_ANX_14_RECOMMENDATION,DISLIST_ROI_RESTRICTION",
                    "_disspact_WAR_disspactportlet_exportDislistsColumns": "DISLIST_CORAP,DISLIST_PBT,DISLIST_DOSSIER_EVALUATION,DISLIST_ED,DISLIST_ARN,DISLIST_ROI_CLH,DISLIST_ROI_SVHC,DISLIST_ANX_14_RECOMMENDATION,DISLIST_ROI_RESTRICTION",
                    "_disspact_WAR_disspactportlet_orderByCol": "name",
                    "_disspact_WAR_disspactportlet_orderByType": "asc",
                    "_disspact_WAR_disspactportlet_orderedSearchableShowListColumns": "DISLIST_PBT_diss_update_date,processes,DISLIST_PBT_diss_concern",
                    "_disspact_WAR_disspactportlet_orderedSearchableShowListElements": "DATE_PICKER,MULTI_VALUE,INPUT_TEXT",
                    "_disspact_WAR_disspactportlet_orderedSearchableShowListProcessColumns": "PACT,PACT,PACT",
                    "_disspact_WAR_disspactportlet_multiValueSearchOperatorprocesses": "AND",
                    "_disspact_WAR_disspactportlet_total": unique_substances,
                    "_disspact_WAR_disspactportlet_exportType": "xls"
                }
            if source == "SVHC":
                dataECHA["_disslists_WAR_disslistsportlet_exportColumns"] = "name,ecNumber,casNumber,haz_detailed_concern,dte_inclusion,doc_cat_decision,doc_cat_iuclid_dossier,doc_cat_supdoc,doc_cat_rcom,prc_external_remarks",
            if source == "SVHCintent":
                dataECHA["_disslists_WAR_disslistsportlet_exportColumns"] = "name,ecNumber,casNumber,sid_other_info_external,sid_avi_index_no,prc_public_status,dte_intention,sbm_expected_submission,sbm_first_submission,dte_withdrawn,lec_submitter,prc_external_remarks,haz_detailed_concern,dte_public_consult_start,dte_public_consult_deadline,doc_cat_report,doc_cat_rcom,prc_msc_agreement_year,doc_cat_agreement,dte_adoption,doc_cat_supdoc,doc_cat_opinion,dte_opinion,doc_cat_minor_opinion,dte_inclusion,diss_update_date,dte_first_published",
            if source == "CoRAP":
                dataECHA["_disslists_WAR_disslistsportlet_exportColumns"] = "name,ecNumber,casNumber,cnt_country,prc_evaluation_year,lec_submitter,haz_detailed_concern,cse_public_lifecycle,diss_update_date,doc_cat_decision,doc_cat_conclusion,doc_cat_justification,dte_corap_publication,lec_contact_address,lec_organization_name,lec_remarks,prc_appeal_link,prc_external_remarks,diss_concern,relevance,dte_first_published",

            responseECHA2 = requests.post(echa_url, params=paramsECHA, data=dataECHA, headers=headersEDass, stream=True)
            if responseECHA2.status_code == 200:
                ECHA_database_bytes = BytesIO(responseECHA2.content)
                logging.info(f"Downloaded {echa_url}")
                responseECHA2.close()
                return ECHA_database_bytes
            else:
                logging.info(f"Failed to download {echa_url}. Status code:", responseECHA2.status_code)
                responseECHA2.close()
                return None
        except requests.exceptions.RequestException as e:
            logging.error(f"Network error while accessing {echa_url}: {e}")
            return None

    #### LOAD DATA SOURCES ####

    # PPP ED
    efsaPPP_url = "https://www.efsa.europa.eu/en/applications/pesticides"
    PPP_ED_string = "overview-endocrine-disrupting-assessment-pesticide-active-substances"
    responseEFSA = requests.get(efsaPPP_url,headers={'User-Agent': random.choice(user_agents_list)})
    ED_PPP = None
    PPP_database_bytes = None
    if responseEFSA.status_code == 200:
        soupEFSA = BeautifulSoup(responseEFSA.text, "html.parser")
        matching_links = [link.get("href") for link in soupEFSA.find_all("a", href=True)
                          if PPP_ED_string in link.get("href") and link.get("href").endswith(('.xls', '.xlsx'))]
        if matching_links:
            file_url = requests.compat.urljoin(efsaPPP_url, matching_links[0])
            ED_PPP = requests.get(file_url)
            if ED_PPP.status_code == 200:
                PPP_database_bytes = BytesIO(ED_PPP.content)
                logging.info("Downloaded EFSA PPP ED database")

    # ED assessment list
    EDass_url = "https://echa.europa.eu/en/ed-assessment"
    EDass_database_bytes = None
    EDass_database_bytes = download_echa_list(EDass_url, user_agents_list,source="EDass")
    # SVHC database
    SVHC_url = "https://echa.europa.eu/en/candidate-list-table"
    SVHC_database_bytes = None
    SVHC_database_bytes = download_echa_list(SVHC_url, user_agents_list,source="SVHC")
    # SVHC intent database
    SVHCintent_url = "https://echa.europa.eu/en/registry-of-svhc-intentions"
    SVHCintent_database_bytes = None
    SVHCintent_database_bytes = download_echa_list(SVHCintent_url, user_agents_list,source="SVHCintent")
    # CoRAP database
    CoRAP_url = "https://echa.europa.eu/en/information-on-chemicals/evaluation/community-rolling-action-plan/corap-table"
    CoRAP_database_bytes = None
    CoRAP_database_bytes = download_echa_list(CoRAP_url, user_agents_list,source="CoRAP")

    # PACT database (function download_echa_list does not work for some reason for PACT)
    PACT_url = "https://echa.europa.eu/en/pact"
    PACT_database_bytes = None
    responsePACT = requests.get(PACT_url,headers={'User-Agent': random.choice(user_agents_list)})
    if responsePACT.status_code == 200:
        soupPACT = BeautifulSoup(responsePACT.text, "html.parser")
        # Find the <small> tag by class
        small_tag = soupPACT.find("small", class_="search-results")
        # Extract the number using regex
        if small_tag:
            text = small_tag.get_text(strip=True)
            match = re.search(r"of\s+([\d,]+)\s+results", text)
            if match:
                unique_substances_PACT = match.group(1).replace(",", "")
        soupPACT.decompose()  # Close/destroy the BS tree
    # Data or payload sent with the POST request+
    paramsPACT = {
        "p_p_id": "disspact_WAR_disspactportlet",
        "p_p_lifecycle": "2",
        "p_p_state": "normal",
        "p_p_mode": "view",
        "p_p_resource_id": "exportResults",
        "p_p_cacheability": "cacheLevelPage"
    }
    dataPACT = {
    "_disspact_WAR_disspactportlet_formDate":int(round(time.time() * 1000)),
    "_disspact_WAR_disspactportlet_exportColumns":"name,ecNumber,casNumber,DISLIST_CORAP,DISLIST_PBT,DISLIST_DOSSIER_EVALUATION,DISLIST_ED,DISLIST_ARN,DISLIST_ROI_CLH,DISLIST_ROI_SVHC,DISLIST_ANX_14_RECOMMENDATION,DISLIST_ROI_RESTRICTION",
    "_disspact_WAR_disspactportlet_exportDislistsColumns":"DISLIST_CORAP,DISLIST_PBT,DISLIST_DOSSIER_EVALUATION,DISLIST_ED,DISLIST_ARN,DISLIST_ROI_CLH,DISLIST_ROI_SVHC,DISLIST_ANX_14_RECOMMENDATION,DISLIST_ROI_RESTRICTION",
    "_disspact_WAR_disspactportlet_orderByCol":"name",
    "_disspact_WAR_disspactportlet_orderByType":"asc",
    "_disspact_WAR_disspactportlet_orderedSearchableShowListColumns":"DISLIST_PBT_diss_update_date,processes,DISLIST_PBT_diss_concern",
    "_disspact_WAR_disspactportlet_orderedSearchableShowListElements":"DATE_PICKER,MULTI_VALUE,INPUT_TEXT",
    "_disspact_WAR_disspactportlet_orderedSearchableShowListProcessColumns":"PACT,PACT,PACT",
    "_disspact_WAR_disspactportlet_multiValueSearchOperatorprocesses":"AND",
    "_disspact_WAR_disspactportlet_total":unique_substances_PACT,
    "_disspact_WAR_disspactportlet_exportType":"xls"
    }
    # Headers (if required)
    headersPACT = {
    "User-Agent": random.choice(user_agents_list),
    "Content-Type": "application/x-www-form-urlencoded",
    }
    # Send the POST request
    responsePACT = requests.post(PACT_url, params=paramsPACT,data=dataPACT, headers=headersPACT, stream=True)
    # Save the database in memory
    if responsePACT.status_code == 200:
        PACT_database_bytes = BytesIO(responsePACT.content)
    else:
        print("Failed to download the PACT list. Status code:", responsePACT.status_code)
    responsePACT.close()

    # BPR ED
    if file_BPR_ED is not None:
        workbookBPR = openpyxl.load_workbook(file_BPR_ED)
    else:
        st.warning("Please upload an Excel file for BPR ED.")
    logging.info(f"BPR ED list loaded successfully")
    # Food additives
    if file_food_add is not None:
        workbook_food_add = openpyxl.load_workbook(file_food_add)
    else:
        st.warning("Please upload an Excel file for food additives.")
    logging.info(f"Food additives list loaded successfully")
    # Food flavourings
    if file_food_flav is not None:
        workbook_food_flav = openpyxl.load_workbook(file_food_flav)
    else:
        st.warning("Please upload an Excel file for food flavourings.")
    logging.info(f"Food flavourings list loaded successfully")

    #### LOOP OVER ALL CAS NUMBERS ####
    i = 0
    while i < len(clp_info):
        CAS = CASall[i]
        clp_info[i]["Input"] = CAS
        st.write(f"Checking chemical: {clp_info[i]["Input"]}")

        # ECHA-CHEM C&L
        clp_info[i]["CAS"] = CAS

        # Check PPP ED list
        if PPP_database_bytes:
            workbook = openpyxl.load_workbook(PPP_database_bytes)
            sheet = workbook.worksheets[0]

            found_PPP_ED = False
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                for cell in row:
                    val = str(cell.value).strip()
                    if val != "-" and val in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):
                        clp_info[i]["ED PPP: Yes/No"] = "Yes"
                        clp_info[i]["ED PPP: Status"] = sheet[f"H{cell.row}"].value
                        clp_info[i]["ED PPP: Conclusion HH"] = sheet[f"I{cell.row}"].value
                        clp_info[i]["ED PPP: Conclusion non-TO"] = sheet[f"J{cell.row}"].value
                        clp_info[i]["ED PPP: EFSA conclusion link"] = sheet[f"N{cell.row}"].value
                        found_PPP_ED = True
                        break
                if found_PPP_ED:
                    break
            if not found_PPP_ED:
                clp_info[i]["ED PPP: Yes/No"] = "No"
        else:
            logging.info("No PPP database")

        # Check ECHA ED Assessment list
        if EDass_database_bytes:
            workbookEDass = openpyxl.load_workbook(EDass_database_bytes)
            # Access the first sheet
            first_sheetEDass = workbookEDass.worksheets[0]
            # Search for a specific string in the first sheet
            found_valueEDass = None
            for rowExcel_EDass in first_sheetEDass.iter_rows(min_row=1, max_row=first_sheetEDass.max_row, values_only=False):
                for cell in rowExcel_EDass:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        EDass_authority = first_sheetEDass[f"E{cell.row}"].value
                        found_valueEDass = EDass_authority
                        EDass_status = first_sheetEDass[f"F{cell.row}"].value
                        EDass_outcome = first_sheetEDass[f"G{cell.row}"].value
                        EDass_lastupdate = first_sheetEDass[f"H{cell.row}"].value
                        clp_info[i]["ED Assessment List: Yes/No"] = "Yes"
                        clp_info[i]["ED Assessment List: Outcome"] = EDass_outcome
                        clp_info[i]["ED Assessment List: Status"] = EDass_status
                        clp_info[i]["ED Assessment List: Authority"] = EDass_authority
                        clp_info[i]["ED Assessment List: Last updated"] = EDass_lastupdate
                        break
                if found_valueEDass:
                    break
            if not found_valueEDass:
                clp_info[i]["ED Assessment List: Yes/No"] = "No"
        else:
            logging.info("No ECHA ED database")

        # Check SVHC list
        if SVHC_database_bytes:
            workbookSVHC = openpyxl.load_workbook(SVHC_database_bytes)
            # Access the first sheet
            first_sheetSVHC = workbookSVHC.worksheets[0]
            # Search for a specific string in the first sheet
            found_valueSVHC = None
            for rowExcel_SVHC in first_sheetSVHC.iter_rows(min_row=1, max_row=first_sheetSVHC.max_row, values_only=False):
                for cell in rowExcel_SVHC:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        SVHCreason = first_sheetSVHC[f"E{cell.row}"].value
                        found_valueSVHC = SVHCreason
                        SVHCdate = first_sheetSVHC[f"I{cell.row}"].value
                        SVHCdecision = first_sheetSVHC[f"J{cell.row}"].value
                        clp_info[i]["SVHC: Yes/No"] = "Yes"
                        clp_info[i]["SVHC: Reason"] = SVHCreason
                        clp_info[i]["SVHC: Date Inclusion"] = SVHCdate
                        clp_info[i]["SVHC: Decision"] = SVHCdecision
                        break
                if found_valueSVHC:
                    break
            if not found_valueSVHC:
                clp_info[i]["SVHC: Yes/No"] = "No"
        else:
            logging.info("No SVHC database")

        # Check SVHC intent list
        if SVHCintent_database_bytes:
            workbookSVHC_intent = openpyxl.load_workbook(SVHCintent_database_bytes)
            # Access the first sheet
            first_sheetSVHC_intent = workbookSVHC_intent.worksheets[0]
            # Search for a specific string in the first sheet
            found_valueSVHC_intent = None
            for rowExcel_SVHC_intent in first_sheetSVHC_intent.iter_rows(min_row=1, max_row=first_sheetSVHC_intent.max_row, values_only=False):
                for cell in rowExcel_SVHC_intent:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        # Extract the value from column G of the same row
                        SVHC_intent_status = first_sheetSVHC_intent[f"G{cell.row}"].value
                        found_valueSVHC_intent = SVHC_intent_status
                        SVHC_intent_scope = first_sheetSVHC_intent[f"N{cell.row}"].value
                        SVHC_intent_lastupdated = first_sheetSVHC_intent[f"AA{cell.row}"].value
                        clp_info[i]["SVHC intent: Yes/No"] = "Yes"
                        clp_info[i]["SVHC intent: Status"] = SVHC_intent_status
                        clp_info[i]["SVHC intent: Scope"] = SVHC_intent_scope
                        clp_info[i]["SVHC intent: Last updated"] = SVHC_intent_lastupdated
                        break
                if found_valueSVHC_intent:
                    break
            if not found_valueSVHC_intent:
                clp_info[i]["SVHC intent: Yes/No"] = "No"
        else:
            logging.info("No SVHC intent database")

        # Check PACT list
        if PACT_database_bytes:
            workbookPACT = openpyxl.load_workbook(PACT_database_bytes)
            # Access the first sheet
            first_sheetPACT = workbookPACT.worksheets[0]
            # Search for a specific string in the first sheet
            found_valuePACT = None
            for rowExcel_PACT in first_sheetPACT.iter_rows(min_row=1, max_row=first_sheetPACT.max_row, values_only=False):
                for cell in rowExcel_PACT:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        # print(f"Found '{CAS}' in row {cell.row}, column {cell.column} on PACT list.")
                        # Extract the SEv from column E of the same row
                        PACT_SEv = first_sheetPACT[f"E{cell.row}"].value
                        found_valuePACT = PACT_SEv
                        PACT_SEv_link = first_sheetPACT[f"F{cell.row}"].value
                        PACT_DEv = first_sheetPACT[f"I{cell.row}"].value
                        PACT_DEv_link = first_sheetPACT[f"J{cell.row}"].value
                        PACT_ED = first_sheetPACT[f"K{cell.row}"].value
                        PACT_ED_link = first_sheetPACT[f"L{cell.row}"].value
                        PACT_ARN = first_sheetPACT[f"M{cell.row}"].value
                        PACT_ARN_link = first_sheetPACT[f"N{cell.row}"].value
                        PACT_PBT = first_sheetPACT[f"G{cell.row}"].value
                        PACT_PBT_link = first_sheetPACT[f"H{cell.row}"].value
                        PACT_CLH = first_sheetPACT[f"O{cell.row}"].value
                        PACT_CLH_link = first_sheetPACT[f"P{cell.row}"].value
                        PACT_SVHC = first_sheetPACT[f"Q{cell.row}"].value
                        PACT_SVHC_link = first_sheetPACT[f"R{cell.row}"].value
                        clp_info[i]["PACT: Yes/No"] = "Yes"
                        clp_info[i]["PACT: SEv"] = PACT_SEv
                        clp_info[i]["PACT: SEv link"] = PACT_SEv_link
                        clp_info[i]["PACT: DEv"] = PACT_DEv
                        clp_info[i]["PACT: DEv link"] = PACT_DEv_link
                        clp_info[i]["PACT: ED"] = PACT_ED
                        clp_info[i]["PACT: ED link"] = PACT_ED_link
                        clp_info[i]["PACT: ARN"] = PACT_ARN
                        clp_info[i]["PACT: ARN link"] = PACT_ARN_link
                        clp_info[i]["PACT: PBT"] = PACT_PBT
                        clp_info[i]["PACT: PBT link"] = PACT_PBT_link
                        clp_info[i]["PACT: CLH"] = PACT_CLH
                        clp_info[i]["PACT: CLH link"] = PACT_CLH_link
                        clp_info[i]["PACT: SVHC"] = PACT_SVHC
                        clp_info[i]["PACT: SVHC link"] = PACT_SVHC_link
                        break
                if found_valuePACT:
                    break
            if not found_valuePACT:
                clp_info[i]["PACT: Yes/No"] = "No"
        else:
            logging.info("No PACT database")

        # Check CoRAP
        if CoRAP_database_bytes:
            workbookCoRAP = openpyxl.load_workbook(CoRAP_database_bytes)
            # Access the first sheet
            first_sheetCoRAP = workbookCoRAP.worksheets[0]
            # Search for a specific string in the first sheet
            found_valueCoRAP = None
            for rowExcel_CoRAP in first_sheetCoRAP.iter_rows(min_row=1, max_row=first_sheetCoRAP.max_row, values_only=False):
                for cell in rowExcel_CoRAP:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]): # Check both CAS or EC
                        CoRAPgrounds = first_sheetCoRAP[f"H{cell.row}"].value
                        found_valueCoRAP = CoRAPgrounds
                        CoRAPstatus = first_sheetCoRAP[f"I{cell.row}"].value
                        CoRAPlastupdate = first_sheetCoRAP[f"J{cell.row}"].value
                        clp_info[i]["CoRAP: Yes/No"] = "Yes"
                        clp_info[i]["CoRAP: Initial grounds of Concern"] = CoRAPgrounds
                        clp_info[i]["CoRAP: Status"] = CoRAPstatus
                        clp_info[i]["CoRAP: Latest update"] = CoRAPlastupdate
                        break
                if found_valueCoRAP:
                    break
            if not found_valueCoRAP:
                clp_info[i]["CoRAP: Yes/No"] = "No"
        else:
            logging.info("No CoRAP database")

        # Check BPR ED list
        if workbookBPR:
            sheetBPR = workbookBPR["List of active substances"]
            # Search for a specific string in the first sheet
            found_valueBPR = None
            for rowExcel_BPR in sheetBPR.iter_rows(min_row=1, max_row=sheetBPR.max_row, values_only=False):
                for cell in rowExcel_BPR:
                    cell_value = str(cell.value).strip().replace("\u00A0", "").replace(" ","")
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        found_valueBPR = sheetBPR[f"K{cell.row}"].value
                        clp_info[i]["BPR: Yes/No"] = "Yes"
                        clp_info[i]["BPR: ED HH"] = sheetBPR[f"K{cell.row}"].value
                        clp_info[i]["BPR: ED ENV"] = sheetBPR[f"L{cell.row}"].value
                        break
                if found_valueBPR:
                    break
            if not found_valueBPR:
                clp_info[i]["BPR: Yes/No"] = "No"
        else:
            logging.info("No BPR ED database")

        # Check Food additives
        if workbook_food_add:
            sheet_food_add = workbook_food_add["List for EDscreener"]
            found_value_food_add = None
            for rowExcel_food_add in sheet_food_add.iter_rows(min_row=1, max_row=sheet_food_add.max_row, values_only=False):
                for cell in rowExcel_food_add:
                    cell_value = str(cell.value).strip()
                    # Based on CAS in column B
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        food_add_Enumber = sheet_food_add[f"B{cell.row}"].value
                        found_value_food_add = food_add_Enumber
                        clp_info[i]["Food additive: Yes/No"] = "Yes"
                        clp_info[i]["Food additive: E number"] = food_add_Enumber
                        break
                    # Also check based on EC in column A (and check if EC is existing)
                    ECnr = clp_info[i]["EC"]
                    if cell.value and ECnr in str(cell.value).strip() and len(ECnr) > 1:
                        food_add_Enumber = sheet_food_add[f"A{cell.row}"].value
                        found_value_food_add = food_add_Enumber
                        clp_info[i]["Food additive: Yes/No"] = "Yes"
                        clp_info[i]["Food additive: E number"] = food_add_Enumber
                        break
                if found_value_food_add:
                    break
            if not found_value_food_add:
                clp_info[i]["Food additive: Yes/No"] = "No"
        else:
            logging.info("No Food additives database")

        #  Food flavourings
        if workbook_food_flav:
            sheet_food_flav = workbook_food_flav["List for EDscreener"]
            # Search for a specific string in the sheet
            found_value_food_flav = None
            for rowExcel_food_flav in sheet_food_flav.iter_rows(min_row=1, max_row=sheet_food_flav.max_row, values_only=False):
                for cell in rowExcel_food_flav:
                    cell_value = str(cell.value).strip()
                    if cell_value != "-" and cell_value in (clp_info[i]["CAS"], clp_info[i]["EC"],clp_info[i]["Input"]):  # Check both CAS or EC
                        found_value_food_flav = "Yes"
                        food_flav_FL = sheet_food_flav[f"A{cell.row}"].value
                        clp_info[i]["Food flavourings: Yes/No"] = "Yes"
                        clp_info[i]["Food flavourings: FL"] = "FL " + food_flav_FL
                        break
                if found_value_food_flav:
                    break
            if not found_value_food_flav:
                clp_info[i]["Food flavourings: Yes/No"] = "No"
        else:
            logging.info("No Food flavourings database")

        # Finalize the loop per chemical
        logging.info(f"Processed {i+1}/{N_CAS}: {clp_info[i]["CAS"]}")
        st.write(f"Processed {i+1}/{N_CAS}: {clp_info[i]["CAS"]}")

        i += 1  # Update in the while loop

    df = pd.DataFrame(clp_info)
    output_excel = BytesIO()
    df.to_excel(output_excel, index=False, engine="openpyxl")
    output_excel.seek(0)


    ### Add Summary sheet to Excel output ####
    # Load workbook
    wb = load_workbook(output_excel)

    # Make the URLs clickable, Loop from the second row to the last row
    ws = wb["Sheet1"]
    columns = [11, 50, 52, 56, 60]  # C&L, PACT SEv, DEv, ARN, CLH
    link_style = Font(color="0000FF", underline="single")
    for row in range(2, ws.max_row + 1):
        for col in columns:
            cell = ws.cell(row=row, column=col)
            cell.hyperlink = cell.value
            cell.font = link_style

    # Check if the "Summary" sheet exists; if not, create it
    if "Summary" not in wb.sheetnames:
        ws = wb.create_sheet(title="Summary")
    else:
        ws = wb["Summary"]

    # Headers to be added in the "Summary" sheet
    headers = [
        ["", "", "", "", "", "Evaluated for ED in", "", "", "", "", "", "Also found in", "", ""],
        ["Name (ECHA-CHEM)", "Input", "CAS number", "EC number", "Classification", "ED assessment",
         "On BPR/PPPR list (for ED-HH; for ED-ENV)", "REACH SVHC candidate", "REACH SVHC intent",
         "CORAP List", "PACT: DEv", "PACT: ARN", "Food lists", "Summary Harmonized", "Summary self-classified"]
    ]
    # Arrange headers and formatting
    for row_index, row_data in enumerate(headers, start=1):
        for col_index, cell_data in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=cell_data)
            # Make the font bold for the header cells
            cell.font = Font(bold=True)
    ws.merge_cells("F1:J1")

    # Insert formulas for summary & Autofill down
    last_row = wb["Sheet1"].max_row
    for row in range(3, last_row + 2):
        ws[f"A{row}"] = f"=Sheet1!B{row - 1}"  # Input
        ws[f"B{row}"] = f"=Sheet1!F{row - 1}"  # Name
        ws[f"C{row}"] = f"=Sheet1!D{row - 1}"  # CAS
        ws[f"D{row}"] = f"=Sheet1!E{row - 1}"  # EC
        ws[f"E{row}"] = f'=IF(Sheet1!N{row - 1}<>"",Sheet1!N{row - 1},"Not classified")'  # Classification
        # ED assessment list
        ws[
            f"F{row}"] = f'=CONCATENATE(Sheet1!AE{row - 1},IF(OR(Sheet1!AE{row - 1}="No",Sheet1!AE{row - 1}=""),""," ("&Sheet1!AF{row - 1}&")"))'
        # BPR/PPP ED?
        ws[
            f"G{row}"] = f'="BPR: "&IF(Sheet1!AB{row - 1}="Yes","Yes (HH: " &Sheet1!AC{row - 1}& "; ENV: " &Sheet1!AD{row - 1}& ")","No")&' \
                         f'"; PPR: "&IF(Sheet1!W{row - 1}="Yes","Yes (HH: "&Sheet1!Y{row - 1} &"; ENV: " &Sheet1!Z{row - 1}& ")","No")'
        # REACH SVHC candidate ED?
        ws[f"H{row}"] = f'=IF(Sheet1!AJ{row - 1}="Yes","Yes: " & Sheet1!AK{row - 1},"No")'
        # REACH SVHC intent ED?
        ws[f"I{row}"] = f'=IF(Sheet1!AR{row - 1}="Yes","Yes: " & Sheet1!AS{row - 1},"No")'
        # CORAP list ED?
        ws[f"J{row}"] = f'=Sheet1!BK{row - 1}&" ("&Sheet1!BL{row - 1}&"; "&Sheet1!BM{row - 1}&")"'
        # PACT: Dev
        ws[f"K{row}"] = f'=Sheet1!AY{row - 1}'
        # PACT: ARN
        ws[f"L{row}"] = f'=Sheet1!BC{row - 1}'
        # Food additives/flavourings
        ws[
            f"M{row}"] = f'=IF(OR(Sheet1!AN{row - 1}="Yes",Sheet1!AP{row - 1}="Yes"),"Yes (" & Sheet1!AO{row - 1} & "; " & Sheet1!AQ{row - 1} & ")", "No")'

    # Determine classification interpretation before adding to Excel
    # Function to determine classification summary
    def determine_classification(value):
        if pd.isna(value) or value.strip() == '':
            return '-'
        for outcome, codes in classification_mapping.items():
            if any(code in value for code in codes):
                return outcome
        return 'other classification'

    # Set up mapping for classification
    classification_mapping = {
        'reproductive toxicity': ['H360', 'H360F', 'H360FD', 'H360Fd', 'H360Df', 'H361', 'H361f', 'H361d', 'H361fd','H362'],
        'STOT-RE': ['H372', 'H373'],
        'carcinogenicity': ['H350', 'H350i', 'H351']
    }

    # Determine classification based on harmonized & self-classified column
    df['Harmonized C&L assessment'] = df['Classification - Hazard statements'].apply(determine_classification)
    df['Self-classified C&L assessment'] = df['Classification - Hazard statements'].apply(determine_classification)
    # If C&L Type is NOT 'Harmonized C&L', set 'Harmonized' column to '-'
    df.loc[df["C&L Type"] != "Harmonised C&L", "Harmonized C&L assessment"] = "-"
    # If C&L Type IS 'Harmonized C&L', set 'Self-class' column to '-'
    df.loc[df["C&L Type"] == "Harmonised C&L", "Self-classified C&L assessment"] = "-"

    # Add to Excel summary sheet
    for i, (val1, val2) in enumerate(zip(df['Harmonized C&L assessment'], df['Self-classified C&L assessment']),
                                     start=3):
        ws[f"N{i}"] = val1
        ws[f"O{i}"] = val2

    # Some formatting
    # Set column width to 12 for columns A to AZ
    for col in range(1, 106):  # 1 to 52 (A to BZ)
        col_letter = get_column_letter(col)
        # For sheet 1
        cell = wb["Sheet1"].cell(row=1, column=col)
        cell.alignment = Alignment(wrap_text=True)
        wb["Sheet1"].column_dimensions[col_letter].width = 12
        # For summary sheet
        for rowN in (1, 2):
            cell = ws.cell(row=rowN, column=col)
            cell.alignment = Alignment(wrap_text=True)
        ws.column_dimensions[col_letter].width = 15

    # Save final workbook
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    ### SAVING ####
    # Save to zip file
    st.session_state.log_stream.seek(0)
    log_bytes = BytesIO(st.session_state.log_stream.read().encode("utf-8"))
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        zip_file.writestr("EDscreener_results.xlsx", final_output.getvalue())
        zip_file.writestr("EDscreener_log.txt", log_bytes.getvalue())
        if PPP_database_bytes:
            zip_file.writestr("databases/EFSA_PPP_ED_Database.xlsx", PPP_database_bytes.getvalue())
        if EDass_database_bytes:
            zip_file.writestr("databases/ED assessment_Database.xlsx", EDass_database_bytes.getvalue())
        if SVHC_database_bytes:
            zip_file.writestr("databases/SVHC_Database.xlsx", SVHC_database_bytes.getvalue())
        if SVHCintent_database_bytes:
            zip_file.writestr("databases/SVHC intent_Database.xlsx", SVHCintent_database_bytes.getvalue())
        if PACT_database_bytes:
            zip_file.writestr("databases/PACT_Database.xlsx", PACT_database_bytes.getvalue())
        if CoRAP_database_bytes:
            zip_file.writestr("databases/CoRAP_Database.xlsx", CoRAP_database_bytes.getvalue())
        if workbookBPR:
            excel_buffer = BytesIO()    # Save workbook to a BytesIO buffer
            workbookBPR.save(excel_buffer)
            excel_buffer.seek(0)  # Reset buffer position to the beginning
            zip_file.writestr(f"databases/{file_BPR_ED.name}", excel_buffer.getvalue())
        if workbook_food_add:
            excel_buffer = BytesIO()    # Save workbook to a BytesIO buffer
            workbook_food_add.save(excel_buffer)
            excel_buffer.seek(0)  # Reset buffer position to the beginning
            zip_file.writestr(f"databases/{file_food_add.name}", excel_buffer.getvalue())
        if workbook_food_flav:
            excel_buffer = BytesIO()    # Save workbook to a BytesIO buffer
            workbook_food_flav.save(excel_buffer)
            excel_buffer.seek(0)  # Reset buffer position to the beginning
            zip_file.writestr(f"databases/{file_food_flav.name}", excel_buffer.getvalue())
    zip_buffer.seek(0)

    return zip_buffer

if uploaded_file:
    if st.button("Run Screener"):
        st.info("Processing started...")
        zip_result = process_data(uploaded_file)
        if zip_result:
            st.download_button("Download All Results (ZIP)", zip_result, file_name=f"EDscreener_package_{datetime.now().strftime("%Y-%m-%d %H-%M")}.zip")
            st.success("Processing finished!")
