
import streamlit as st
import pandas as pd
import re
import requests
from bs4 import BeautifulSoup
from io import BytesIO, StringIO
import openpyxl
from datetime import datetime
import logging
import zipfile

st.title("ED Screener Tool")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

# In-memory log stream
log_stream = StringIO()
logging.basicConfig(stream=log_stream, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def process_data(file):
    logging.info("Started ED screener process")
    CASallpd = pd.read_excel(file, engine="openpyxl")
    if "CAS" not in CASallpd.columns:
        st.error("Error: 'CAS' column not found.")
        return None

    CASall = CASallpd["CAS"].dropna().tolist()
    CASall = [re.sub(r'[^\d\-]', '', str(cas)) for cas in CASall]
    N_CAS = len(CASall)

    clp_info = [{"id": i+1, "Input": CASall[i]} for i in range(N_CAS)]
    key_names = ["Input", "ED PPP: Yes/No", "ED PPP: Status", "ED PPP: Conclusion HH", "ED PPP: Conclusion non-TO", "ED PPP: EFSA conclusion link"]
    for entry in clp_info:
        for key in key_names:
            entry[key] = "-"

    efsaPPP_url = "https://www.efsa.europa.eu/en/applications/pesticides"
    PPP_ED_string = "overview-endocrine-disrupting-assessment-pesticide-active-substances"
    responseEFSA = requests.get(efsaPPP_url)
    ED_PPP = None
    database_bytes = None

    if responseEFSA.status_code == 200:
        soupEFSA = BeautifulSoup(responseEFSA.text, "html.parser")
        matching_links = [link.get("href") for link in soupEFSA.find_all("a", href=True)
                          if PPP_ED_string in link.get("href") and link.get("href").endswith(('.xls', '.xlsx'))]
        if matching_links:
            file_url = requests.compat.urljoin(efsaPPP_url, matching_links[0])
            ED_PPP = requests.get(file_url)
            if ED_PPP.status_code == 200:
                database_bytes = BytesIO(ED_PPP.content)
                logging.info("Downloaded EFSA PPP ED database")

    if ED_PPP and ED_PPP.status_code == 200:
        workbook = openpyxl.load_workbook(database_bytes)
        sheet = workbook.worksheets[0]

        for i, entry in enumerate(clp_info):
            found = False
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                for cell in row:
                    val = str(cell.value).strip()
                    if val == entry["Input"]:
                        entry["ED PPP: Yes/No"] = "Yes"
                        entry["ED PPP: Status"] = sheet[f"H{cell.row}"].value
                        entry["ED PPP: Conclusion HH"] = sheet[f"I{cell.row}"].value
                        entry["ED PPP: Conclusion non-TO"] = sheet[f"J{cell.row}"].value
                        entry["ED PPP: EFSA conclusion link"] = sheet[f"N{cell.row}"].value
                        found = True
                        break
                if found:
                    break
            logging.info(f"Processed {i+1}/{N_CAS}")
            st.write(f"Processed {i+1}/{N_CAS}")

    df = pd.DataFrame(clp_info)
    output_excel = BytesIO()
    df.to_excel(output_excel, index=False, engine="openpyxl")
    output_excel.seek(0)

    log_stream.seek(0)
    log_bytes = BytesIO(log_stream.read().encode("utf-8"))

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        zip_file.writestr("EDscreener_results.xlsx", output_excel.getvalue())
        zip_file.writestr("EDscreener_log.txt", log_bytes.getvalue())
        if database_bytes:
            zip_file.writestr("EFSA_PPP_ED_Database.xlsx", database_bytes.getvalue())
    zip_buffer.seek(0)

    return zip_buffer

if uploaded_file:
    if st.button("Run Screener"):
        st.info("Processing started...")
        zip_result = process_data(uploaded_file)
        if zip_result:
            st.download_button("Download All Results (ZIP)", zip_result, file_name="EDscreener_package.zip")
            st.success("Processing finished!")
